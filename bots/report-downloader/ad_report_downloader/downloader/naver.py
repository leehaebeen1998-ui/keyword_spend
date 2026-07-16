"""Naver 검색광고 다차원 보고서 다운로더 — ads.naver.com
URL: https://ads.naver.com/manage/ad-accounts/{account_id}/sa/reports
날짜 형식: YYYY.MM.DD. (점 포함)
"""
from __future__ import annotations
import re
import tempfile
from datetime import date, timedelta
from pathlib import Path
from typing import TYPE_CHECKING

import yaml

if TYPE_CHECKING:
    from playwright.sync_api import Page

from downloader.base import BaseDownloader, EmptyDataError

_SELECTOR_FILE = Path(__file__).parent.parent / "selectors" / "naver.yaml"


def _load_selectors() -> dict:
    with _SELECTOR_FILE.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


class NaverDownloader(BaseDownloader):
    MEDIA_CODE = "naver"
    IMPLEMENTED = True

    def __init__(self, config: dict, account: dict | None = None):
        super().__init__(config, account)
        self._sel = _load_selectors()
        # account_id는 BaseDownloader.__init__에서 self.account_id로 설정됨

    def _report_list_url(self) -> str:
        """Return the full URL of the dimension report list page."""
        path = self._sel["report"]["report_list_path"].replace(
            "{account_id}", self.account_id
        )
        return self._sel["base_url"] + path

    # ------------------------------------------------------------------ #
    # 1. Login check                                                       #
    # ------------------------------------------------------------------ #
    def check_login(self, page: "Page") -> bool:
        target_url = self._report_list_url() if self.account_id else self._sel["base_url"]
        try:
            page.goto(target_url, wait_until="networkidle", timeout=40_000)
        except Exception:
            # networkidle 타임아웃 허용 — URL 체크로 로그인 여부 판단
            pass
        self.human_delay(500, 1000)

        url = page.url.lower()
        for pattern in self._sel["login"]["login_url_patterns"]:
            if pattern.lower() in url:
                self._info(f"로그인 필요 (URL: {page.url[:80]})")
                return False
        try:
            if page.locator(self._sel["login"]["login_element_selector"]).first.is_visible(timeout=2_000):
                self._info("로그인 필요 (로그인 폼 감지)")
                return False
        except Exception:
            pass
        return True

    # ------------------------------------------------------------------ #
    # 2. Navigate to the saved report                                      #
    # ------------------------------------------------------------------ #
    def navigate_to_report(self, page: "Page") -> None:
        sel = self._sel["report"]

        if self.account_id:
            # check_login이 이미 이동했으면 재이동 생략, 아니면 이동
            current = page.url.lower()
            target = self._report_list_url().lower()
            if target not in current:
                self._info(f"보고서 목록 이동 (account: {self.account_id})")
                try:
                    page.goto(self._report_list_url(), wait_until="networkidle", timeout=30_000)
                except Exception:
                    pass
            else:
                self._info(f"보고서 목록 이미 로드됨 (account: {self.account_id})")

            # 이동 후 로그인 페이지인지 재확인
            url_after = page.url.lower()
            for pattern in self._sel["login"]["login_url_patterns"]:
                if pattern.lower() in url_after:
                    from downloader.base import LoginRequiredError
                    raise LoginRequiredError()
            self.human_delay(800, 1500)
        else:
            self._warn("account_id 미설정 — 사이드바 네비게이션 시도 (여러 계정 시 실패 가능)")
            try:
                section = page.locator(sel["sidebar_report"]).first
                if section.is_visible(timeout=3_000):
                    section.click()
                    self.human_delay(300, 600)
            except Exception:
                pass
            try:
                dim = page.locator(sel["sidebar_multidim"]).first
                dim.wait_for(state="visible", timeout=8_000)
                dim.click()
                self.human_delay(800, 1500)
                page.wait_for_load_state("networkidle", timeout=20_000)
            except Exception as e:
                raise RuntimeError("보고서 목록 이동 실패: account_id 미설정") from e

        self._click_report_by_name(page)

    def _click_report_by_name(self, page: "Page") -> None:
        name = self.report_name
        self._info(f"보고서 찾는 중: '{name}'")

        # 페이지네이션 포함 최대 10페이지까지 탐색
        for page_num in range(1, 11):
            if page_num > 1:
                self._info(f"보고서 목록 {page_num}페이지 탐색 중...")

            # 방법 1: exact text match
            try:
                target = page.get_by_text(name, exact=True).first
                target.wait_for(state="visible", timeout=5_000)
                target.click()
                self.human_delay(800, 1500)
                page.wait_for_load_state("networkidle", timeout=25_000)
                self._info(f"보고서 '{name}' 열림 (페이지 {page_num})")
                return
            except Exception:
                pass

            # 방법 2: href 기반 링크 순회 (정확히 일치하는 이름만 클릭)
            #
            # 주의: 여기서 부분 문자열(in) 매칭을 쓰면 안 된다. 예를 들어
            # "오현 키워드 소진액"은 "오현 키워드 소진액(스프레드)"의 부분
            # 문자열이라서, 부분 매칭을 쓰면 정확한 이름을 요청했는데도
            # "(스프레드)"가 붙은 다른 보고서를 잘못 클릭하게 된다.
            # (공백/개행 차이만 흡수하기 위해 whitespace만 정규화한다.)
            try:
                links = page.locator(self._sel["report"]["report_list_link"]).all()
                for link in links:
                    try:
                        text = " ".join(link.inner_text(timeout=1_000).split())
                        if text == name:
                            link.click()
                            self.human_delay(800, 1500)
                            page.wait_for_load_state("networkidle", timeout=25_000)
                            self._info(f"보고서 '{name}' 열림 (fallback, 페이지 {page_num})")
                            return
                    except Exception:
                        continue
            except Exception:
                pass

            # 현재 페이지에서 못 찾음 — 다음 페이지 버튼 클릭 시도
            next_btn = self._find_next_page_button(page)
            if next_btn is None:
                self._info(f"다음 페이지 없음 (마지막 페이지: {page_num})")
                break
            next_btn.click()
            self.human_delay(1000, 1500)
            try:
                page.wait_for_load_state("networkidle", timeout=15_000)
            except Exception:
                pass

        self._screenshot(page, "naver_report_list")
        raise RuntimeError(
            f"보고서 '{name}' 없음. "
            "config.json → media.naver.accounts[].report_name 확인"
        )

    def _find_next_page_button(self, page: "Page"):
        """보고서 목록의 다음 페이지 버튼을 찾아 반환. 없으면 None.

        네이버 광고 UI 페이지네이션 구조:
          <li title="다음 페이지" class="ad-cms-pagination-next ..." aria-disabled="false">
        — button이 아닌 li 요소임에 주의
        """
        candidates = [
            # 네이버 광고 UI: li 기반 페이지네이션
            "li.ad-cms-pagination-next:not([aria-disabled='true'])",
            "li[title='다음 페이지']:not([aria-disabled='true'])",
            # fallback
            "[class*='pagination-next']:not([aria-disabled='true'])",
            "li[title='다음']:not([aria-disabled='true'])",
            # 숫자 페이지 버튼 방식
            "button[aria-label='다음 페이지']:not(:disabled)",
            "button[title='다음 페이지']:not(:disabled)",
        ]
        for sel in candidates:
            try:
                btn = page.locator(sel).first
                if btn.is_visible(timeout=1_000):
                    return btn
            except Exception:
                continue
        return None

    # ------------------------------------------------------------------ #
    # 3. Set date period                                                   #
    # ------------------------------------------------------------------ #
    def set_period(self, page: "Page", start: date, end: date) -> None:
        start_str = start.strftime("%Y.%m.%d.")
        end_str   = end.strftime("%Y.%m.%d.")
        sel = self._sel["period"]

        self._info(f"날짜 피커 열기: {start_str} ~ {end_str}")

        # ── 1) 날짜 버튼 클릭 ──────────────────────────────────────────
        opened = self._open_date_picker(page, sel, start_str)
        if not opened:
            self._warn("날짜 피커를 열 수 없음 — 날짜 설정 건너뜀")
            self._screenshot(page, "naver_date_picker_not_found")
            return

        # ── 2) 팝업 렌더링 대기 ────────────────────────────────────────
        self.human_delay(600, 1000)
        self._screenshot(page, "naver_date_picker_opened")  # 피커 열린 직후 캡처

        # ── 3) 캘린더 날짜 직접 클릭 ─────────────────────────────────
        # rwc-month[data-year][data-month] li[data-day] button 구조 사용
        if not self._pick_dates_by_calendar(page, start, end):
            self._warn("캘린더 날짜 클릭 실패 — 날짜 설정 건너뜀")
            return

        # ── 5) 확인 버튼 ───────────────────────────────────────────────
        try:
            confirm = page.locator(sel["confirm_button"]).first
            confirm.wait_for(state="visible", timeout=4_000)
            confirm.click()
            self.human_delay(600, 1000)
        except Exception:
            page.keyboard.press("Enter")
            self.human_delay(600, 1000)

        page.wait_for_load_state("networkidle", timeout=25_000)
        self._info(f"날짜 설정 완료: {start_str} ~ {end_str}")

    def _open_date_picker(self, page: "Page", sel: dict, start_str: str) -> bool:
        """날짜 범위 버튼을 찾아 클릭. 성공 시 True 반환."""
        # 방법 1: YAML에 정의된 셀렉터 순서대로 시도
        for trigger_sel in sel["date_range_button"].split(","):
            trigger_sel = trigger_sel.strip()
            if not trigger_sel:
                continue
            try:
                btn = page.locator(trigger_sel).first
                if btn.is_visible(timeout=1_500):
                    btn.click()
                    return True
            except Exception:
                continue

        # 방법 2: 현재 날짜 텍스트가 포함된 span의 부모 버튼 클릭
        # (날짜 형식: YYYY.MM.DD. — 연월 부분으로 탐색)
        month_str = start_str[:7]  # "2026.06"
        try:
            # span 중 날짜 텍스트를 포함하는 것을 찾아 클릭
            spans = page.locator(f"span:text-matches('{month_str}', 'i')").all()
            for sp in spans[:3]:
                try:
                    if sp.is_visible(timeout=500):
                        sp.click()
                        return True
                except Exception:
                    continue
        except Exception:
            pass

        return False

    def _pick_dates_by_calendar(self, page: "Page", start: date, end: date) -> bool:
        """
        rwc-month 달력에서 날짜 직접 클릭.
        구조: [data-year=YYYY][data-month=M(0-indexed)] li[data-day=D] button
        """
        def _click_day(target: date) -> bool:
            month_0 = target.month - 1   # 0-indexed
            # 특정 년/월의 특정 일 버튼
            sel = (
                f'[data-year="{target.year}"][data-month="{month_0}"] '
                f'li[data-day="{target.day}"] button'
            )
            try:
                btn = page.locator(sel).first
                btn.wait_for(state="visible", timeout=5_000)
                btn.click()
                self.human_delay(300, 500)
                self._info(f"캘린더 클릭: {target}")
                return True
            except Exception as e:
                self._warn(f"캘린더 클릭 실패 ({target}): {e}")
                return False

        ok_start = _click_day(start)
        self.human_delay(300, 500)
        ok_end   = _click_day(end)
        return ok_start and ok_end

    def _wait_for_date_inputs(self, page: "Page", sel: dict) -> bool:
        """날짜 텍스트 입력 필드가 나타날 때까지 최대 5초 대기."""
        candidates = [
            "input.ad-cms-input[placeholder='YYYY.MM.DD.']",
            "input[placeholder*='YYYY.MM.DD']",
            "input[placeholder*='날짜']",
        ]
        # YAML start_input 셀렉터도 추가
        for part in sel.get("start_input", "").split(","):
            part = part.strip()
            if part:
                candidates.append(part)

        for s in candidates:
            try:
                page.locator(s).first.wait_for(state="visible", timeout=2_000)
                return True
            except Exception:
                continue
        return False

    def _fill_date_range(self, page: "Page", start_str: str, end_str: str) -> bool:
        """
        날짜 피커 내 두 개의 input에 시작/종료일을 입력.
        React 컨트롤드 인풋 대응: native value setter + 이벤트 dispatch.
        """
        sel = "input.ad-cms-input[placeholder='YYYY.MM.DD.']"
        try:
            inputs = page.locator(sel).all()
        except Exception:
            inputs = []

        if len(inputs) < 2:
            self._warn(f"날짜 입력창 {len(inputs)}개만 발견 (2개 필요)")
            return False

        # React controlled input에 값을 강제 설정하는 JS
        _set_value_js = """
            (el, value) => {
                const setter = Object.getOwnPropertyDescriptor(
                    window.HTMLInputElement.prototype, 'value').set;
                setter.call(el, value);
                el.dispatchEvent(new Event('input',  { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
            }
        """

        def _react_fill(inp, value: str) -> bool:
            try:
                inp.click()
                self.human_delay(80, 120)
                inp.evaluate(_set_value_js, value)
                self.human_delay(150, 250)
                return True
            except Exception as e:
                self._warn(f"React fill 실패 ({value}): {e}")
                # fallback: 직접 타이핑
                try:
                    inp.triple_click()
                    page.keyboard.type(value, delay=40)
                    self.human_delay(150, 250)
                    return True
                except Exception:
                    return False

        ok_start = _react_fill(inputs[0], start_str)
        self.human_delay(200, 300)
        ok_end   = _react_fill(inputs[1], end_str)
        self.human_delay(200, 300)

        if ok_start and ok_end:
            self._info(f"날짜 입력 완료: {start_str} ~ {end_str}")
        return ok_start and ok_end

    def _fill_date_input(self, page: "Page", selector: str, value: str) -> None:
        """React 컨트롤드 인풋 — keyboard.type()으로 실제 키 이벤트 발생."""
        # 가장 신뢰도 높은 셀렉터만 사용 (placeholder 기반)
        candidates = [
            "input.ad-cms-input[placeholder='YYYY.MM.DD.']",
            "input[placeholder='YYYY.MM.DD.']",
        ]
        # YAML 셀렉터도 추가
        for part in selector.split(","):
            part = part.strip()
            if part and part not in candidates:
                candidates.append(part)

        for s in candidates:
            try:
                # start_input이면 first, end_input이면 nth(1)을 호출하는 쪽에서 결정
                # 여기서는 locator를 반환하는 것을 쓰지 않고 직접 처리
                inp = page.locator(s).first
                if not inp.is_visible(timeout=1_500):
                    continue
                inp.click()
                self.human_delay(100, 200)
                # 전체 선택 후 새 값 입력 (keyboard.type = 실제 키 이벤트)
                page.keyboard.key("Control+a")
                self.human_delay(50, 100)
                page.keyboard.type(value)
                self.human_delay(200, 300)
                return
            except Exception:
                continue
        self._warn(f"날짜 입력 실패: {value}")

    # ------------------------------------------------------------------ #
    # 4. Trigger download                                                  #
    # ------------------------------------------------------------------ #
    def trigger_download(self, page: "Page", start: date, end: date) -> Path:
        # ── 데이터 초과 에러 체크 ─────────────────────────────────────────
        if self._is_data_too_large(page):
            self._warn(f"데이터 초과 ({start}~{end}) — 기간 분할 다운로드")
            return self._download_split(page, start, end)

        dl_cfg = self._sel["download"]
        timeout_ms = self.timeout_sec * 1_000

        btn = self._find_download_button(page, dl_cfg)
        if btn is None:
            self._screenshot(page, "naver_download_btn_not_found")
            raise RuntimeError("다운로드 버튼 없음. selectors/naver.yaml 확인")

        if btn.is_disabled():
            raise EmptyDataError("다운로드 버튼 비활성 — 해당 기간 데이터 없음")

        self.human_delay(300, 700)

        with page.expect_download(timeout=timeout_ms) as dl_info:
            btn.click()
            # 드롭다운/모달이 뜨는 경우 처리
            self._handle_download_modal(page)

        download = dl_info.value
        suggested = download.suggested_filename or "naver_report.xlsx"
        suffix = Path(suggested).suffix or ".xlsx"
        tmp = Path(tempfile.mktemp(suffix=suffix))
        download.save_as(str(tmp))

        if not tmp.exists() or tmp.stat().st_size == 0:
            raise RuntimeError(f"다운로드 파일 비어있음: {tmp}")

        self._info(f"파일: {suggested} ({tmp.stat().st_size:,} bytes)")
        return tmp

    def _find_download_button(self, page: "Page", dl_cfg: dict):
        for key in ("button", "button_fallback"):
            sel = dl_cfg.get(key, "")
            if not sel:
                continue
            for s in sel.split(","):
                s = s.strip()
                if not s:
                    continue
                try:
                    loc = page.locator(s).first
                    loc.wait_for(state="visible", timeout=8_000)
                    return loc
                except Exception:
                    continue
        return None

    def _handle_download_modal(self, page: "Page") -> None:
        """다운로드 버튼 클릭 후 드롭다운/모달이 뜨면 추가 클릭 처리.

        네이버 다차원 보고서의 다운로드 버튼은 클릭 시
        드롭다운 메뉴(Excel/CSV 선택) 또는 확인 모달이 뜰 수 있음.
        """
        self.human_delay(500, 800)
        self._screenshot(page, "naver_after_download_click")

        # 드롭다운: 엑셀/Excel 항목 우선 클릭
        dropdown_candidates = [
            "li:has-text('엑셀')",
            "li:has-text('Excel')",
            "li:has-text('.xlsx')",
            "[role='menuitem']:has-text('엑셀')",
            "[role='menuitem']:has-text('Excel')",
            "[role='option']:has-text('엑셀')",
            "button:has-text('엑셀'):not([class*='btn-variant-text'])",
            "li:has-text('CSV')",
            "[role='menuitem']:has-text('CSV')",
        ]
        for sel in dropdown_candidates:
            try:
                item = page.locator(sel).first
                if item.is_visible(timeout=500):   # 짧게 — 없으면 바로 넘김
                    self._info(f"다운로드 포맷 선택: {sel}")
                    item.click()
                    return
            except Exception:
                continue

        # 모달 확인 버튼 (다운로드 확인 팝업)
        modal_candidates = [
            "button.ad-cms-btn-color-primary:has-text('다운로드')",
            "button.ad-cms-btn-color-primary:has-text('확인')",
            "[role='dialog'] button:has-text('다운로드')",
            "[role='dialog'] button:has-text('확인')",
        ]
        for sel in modal_candidates:
            try:
                item = page.locator(sel).first
                if item.is_visible(timeout=500):   # 짧게 — 없으면 바로 넘김
                    self._info(f"다운로드 모달 확인: {sel}")
                    item.click()
                    return
            except Exception:
                continue

        # 드롭다운/모달 없음 → 버튼 클릭 자체가 다운로드 트리거
        self._info("다운로드 포맷 선택 불필요 (직접 다운로드)")

    # ------------------------------------------------------------------ #
    # 5. 데이터 초과 시 분할 다운로드                                      #
    # ------------------------------------------------------------------ #

    def _is_data_too_large(self, page: "Page") -> bool:
        """'데이터양이 지나치게 많아' 에러가 페이지에 보이면 True."""
        try:
            loc = page.locator("text=데이터양이 지나치게 많아").first
            return loc.is_visible(timeout=1_500)
        except Exception:
            return False

    def _download_split(self, page: "Page", start: date, end: date) -> Path:
        """날짜 범위를 절반으로 분할하여 각각 다운로드 후 병합."""
        delta = (end - start).days + 1
        if delta <= 1:
            self._screenshot(page, "naver_data_too_large_1day")
            raise RuntimeError(
                f"1일 단위에서도 데이터 초과: {start}. "
                "보고서 항목을 줄이거나 네이버에서 직접 다운로드 필요"
            )

        mid = start + timedelta(days=delta // 2 - 1)
        ranges = [(start, mid), (mid + timedelta(days=1), end)]
        chunk_paths: list[Path] = []

        for chunk_start, chunk_end in ranges:
            self._info(f"  청크: {chunk_start} ~ {chunk_end}")
            try:
                page.goto(self._report_list_url(), wait_until="networkidle", timeout=30_000)
            except Exception:
                pass
            self.human_delay(500, 1000)
            self._click_report_by_name(page)
            self.set_period(page, chunk_start, chunk_end)
            # trigger_download 재귀 — 청크가 여전히 크면 다시 분할
            tmp = self.trigger_download(page, chunk_start, chunk_end)
            chunk_paths.append(tmp)

        return self._merge_files(chunk_paths)

    def _merge_files(self, paths: list[Path]) -> Path:
        """여러 청크 파일을 하나로 합침 (xlsx / csv 자동 판별)."""
        if not paths:
            raise RuntimeError("병합할 파일 없음")
        if len(paths) == 1:
            return paths[0]

        suffix = paths[0].suffix.lower()
        if suffix == ".xlsx":
            return self._merge_xlsx(paths)
        elif suffix in (".csv", ".tsv"):
            return self._merge_csv(paths, suffix)
        else:
            self._warn(f"알 수 없는 형식 ({suffix}) — 첫 번째 파일만 사용")
            return paths[0]

    def _merge_xlsx(self, paths: list[Path]) -> Path:
        """openpyxl로 xlsx 청크 병합."""
        try:
            import openpyxl
        except ImportError:
            self._warn("openpyxl 미설치 — 첫 번째 청크만 사용")
            return paths[0]

        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        header_written = False

        for p in paths:
            try:
                wb_in = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                ws_in = wb_in.active
                first_row = True
                for row in ws_in.iter_rows(values_only=True):
                    if first_row:
                        first_row = False
                        if not header_written:
                            ws_out.append(list(row))
                            header_written = True
                        # else: 중복 헤더 건너뜀
                    else:
                        ws_out.append(list(row))
                wb_in.close()
            except Exception as e:
                self._warn(f"xlsx 병합 오류 ({p.name}): {e}")

        out = Path(tempfile.mktemp(suffix=".xlsx"))
        wb_out.save(str(out))
        self._info(f"xlsx {len(paths)}개 병합 완료")
        return out

    def _merge_csv(self, paths: list[Path], suffix: str = ".csv") -> Path:
        """CSV/TSV 청크 병합 (헤더는 첫 번째만)."""
        import csv

        out = Path(tempfile.mktemp(suffix=suffix))
        with out.open("w", newline="", encoding="utf-8-sig") as f_out:
            writer = csv.writer(f_out)
            header_written = False
            for p in paths:
                try:
                    with p.open("r", encoding="utf-8-sig") as f_in:
                        for i, row in enumerate(csv.reader(f_in)):
                            if i == 0:
                                if not header_written:
                                    writer.writerow(row)
                                    header_written = True
                            else:
                                writer.writerow(row)
                except Exception as e:
                    self._warn(f"csv 병합 오류 ({p.name}): {e}")

        self._info(f"csv {len(paths)}개 병합 완료")
        return out

