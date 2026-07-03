import unittest

from pathlib import Path
from tempfile import TemporaryDirectory

from index_classifier.brand_upload import (
    DEFAULT_BRAND_RULES,
    build_sheet_targets,
    offset_for_report_date,
    report_date_for_offset,
    target_dates_for_upload,
    weekend_catchup_offsets,
)
from index_classifier.brand_template_writer import write_brand_template
from index_classifier.raw_upload_builder import build_upload_rows_from_raw, sort_upload_rows


class BrandUploadTests(unittest.TestCase):
    def test_today_minus_one_uses_previous_calendar_day(self):
        self.assertEqual(str(report_date_for_offset("2026-07-01", 1)), "2026-06-30")
        self.assertEqual(offset_for_report_date("2026-07-01", "2026-06-30"), 1)

    def test_ohyun_fixed_targets_match_one_day_sheets(self):
        targets = build_sheet_targets(DEFAULT_BRAND_RULES["법무법인 오현"], "2026-07-01")

        self.assertIn("ohcrime(형사)1일", [target.sheet_name for target in targets])
        self.assertIn("파컨(교통사고)1일", [target.sheet_name for target in targets])
        self.assertIn("파컨(이혼)1일", [target.sheet_name for target in targets])
        self.assertIn("파컨(경제범죄)1일", [target.sheet_name for target in targets])
        self.assertTrue(all(str(target.report_date) == "2026-06-30" for target in targets))
        self.assertTrue(all(target.date_formula == "=TODAY()-1" for target in targets))

    def test_taeha_rolling_targets_include_naver_and_google(self):
        targets = build_sheet_targets(DEFAULT_BRAND_RULES["법무법인 태하"], "2026-07-01")
        sheet_names = {target.sheet_name for target in targets}

        self.assertIn("형(1일)", sheet_names)
        self.assertIn("형(7일)", sheet_names)
        self.assertIn("성범죄(1일)_구글", sheet_names)
        self.assertIn("군형사(7일)_구글", sheet_names)

    def test_monday_run_derives_weekend_catchup_dates(self):
        self.assertEqual(weekend_catchup_offsets("2026-07-06"), (1, 2, 3))
        self.assertEqual(
            target_dates_for_upload("2026-07-06", weekend_catchup_offsets("2026-07-06")),
            ("20260705", "20260704", "20260703"),
        )

    def test_ohyun_xlsx_template_writer_updates_matching_sheet(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "output.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "ohcrime(형사)1일"
            worksheet.append([])
            worksheet.append(["", "키워드별 소진액 보고서_http://www.ohcrime.com_형사", "", "", "", "", "", "", "=TODAY()-1"])
            worksheet.append([])
            worksheet.append(["캠페인유형", "PC/모바일", "키워드", "노출수", "클릭수", "클릭률", "클릭비용", "총비용", "노출순위"])
            worksheet.append([])
            worksheet.append(["파워링크", "PC", "old", 1, 1, 1, 1, 1, 1])
            workbook.save(template)

            result = write_brand_template(
                brand="법무법인 오현",
                template_path=template,
                output_path=output,
                run_date="2026-07-01",
                rows=[
                    {
                        "date": "20260630",
                        "media": "Naver",
                        "category": "형사",
                        "campaign_type": "파워링크",
                        "device": "모바일",
                        "keyword": "형사변호사",
                        "impressions": "100",
                        "clicks": "5",
                        "cost": "50000",
                        "rank": "2.4",
                    }
                ],
            )

            self.assertEqual(result.written_rows, 1)
            self.assertEqual(result.touched_sheets, ["ohcrime(형사)1일"])
            saved = load_workbook(output, data_only=False)
            sheet = saved["ohcrime(형사)1일"]
            self.assertEqual(sheet["A6"].value, "파워링크")
            self.assertEqual(sheet["C6"].value, "형사변호사")
            self.assertEqual(sheet["F6"].value, 0.05)
            self.assertEqual(sheet["G6"].value, 10000)
            self.assertEqual(sheet["H6"].value, 50000)
            self.assertEqual(sheet["I6"].value, 2.4)

    def test_ohyun_splits_powerlink_and_powercontents_and_skips_zero_cost(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "output.xlsx"
            workbook = Workbook()
            workbook.remove(workbook.active)
            for sheet_name in ("교통사고1일", "파컨(교통사고)1일"):
                worksheet = workbook.create_sheet(sheet_name)
                worksheet.append([])
                worksheet.append([])
                worksheet.append([])
                worksheet.append(["캠페인유형", "PC/모바일", "키워드", "노출수", "클릭수", "클릭률", "클릭비용", "총비용", "노출순위"])
                worksheet.append([])
            workbook.save(template)

            result = write_brand_template(
                brand="법무법인 오현",
                template_path=template,
                output_path=output,
                run_date="2026-07-01",
                rows=[
                    {"date": "20260630", "media": "Naver", "category": "교통사고", "campaign_type": "파워링크", "keyword": "교통 링크", "cost": "300"},
                    {"date": "20260630", "media": "Naver", "category": "교통사고", "campaign_type": "파워컨텐츠", "keyword": "교통 컨텐츠", "cost": "500"},
                    {"date": "20260630", "media": "Naver", "category": "교통사고", "campaign_type": "파워링크", "keyword": "제로", "cost": "0"},
                ],
            )

            self.assertEqual(result.written_rows, 2)
            saved = load_workbook(output, data_only=False)
            self.assertEqual(saved["교통사고1일"]["C6"].value, "교통 링크")
            self.assertEqual(saved["파컨(교통사고)1일"]["C6"].value, "교통 컨텐츠")

    def test_template_writer_forces_blank_conversion_metrics_to_zero(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "output.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "형(1일)"
            worksheet.append(["키워드", "총비용", "전환수", "전환율", "전환당비용"])
            worksheet.append([])
            workbook.save(template)

            result = write_brand_template(
                brand="법무법인 태하",
                template_path=template,
                output_path=output,
                run_date="2026-07-01",
                rows=[{"date": "20260630", "media": "Naver", "category": "형", "keyword": "형사", "cost": "100"}],
            )

            self.assertEqual(result.written_rows, 1)
            saved = load_workbook(output, data_only=False)
            sheet = saved["형(1일)"]
            self.assertEqual(sheet["C3"].value, 0)
            self.assertEqual(sheet["D3"].value, 0)
            self.assertEqual(sheet["E3"].value, 0)
            self.assertEqual(sheet["C3"].number_format, "0;-0;-")
            self.assertEqual(sheet["D3"].number_format, "0.00%;-0.00%;-")
            self.assertEqual(sheet["E3"].number_format, "#,##0;-#,##0;-")

    def test_taeha_naver_account_number_overrides_keyword_guess(self):
        with TemporaryDirectory() as temp_dir:
            rules = Path(temp_dir) / "rules.csv"
            rules.write_text(
                "브랜드,순위,규칙,매칭값,카테고리,신뢰도,사용,메모\n"
                "법무법인 태하,0,계정번호,01,형,1,O,\n",
                encoding="utf-8-sig",
            )
            raw = Path(temp_dir) / "naver_thlaw_01_raw_20260630_20260630.csv"
            raw.write_text(
                '"일일보고서_0920(2026.06.30.~2026.06.30.),1826631"\n'
                "캠페인유형,PC/모바일 매체,키워드,노출수,클릭수,클릭률(%),평균 CPC,총비용,총 전환수,총 전환율(%),총 전환당비용(원),평균노출순위\n"
                "파워링크,모바일,교통사고변호사,100,1,1,5000,5000,0,0,0,2.1\n",
                encoding="utf-8-sig",
            )

            rows = build_upload_rows_from_raw(brand="법무법인 태하", media="Naver", input_path=raw, rules_path=rules)

            self.assertEqual(rows[0]["category"], "형")
            self.assertEqual(rows[0]["date"], "20260630")

    def test_template_writer_does_not_duplicate_undated_rows_across_rolling_sheets(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "output.xlsx"
            workbook = Workbook()
            workbook.remove(workbook.active)
            for sheet_name in ("형(1일)", "형(2일)"):
                worksheet = workbook.create_sheet(sheet_name)
                worksheet.append(["키워드", "총비용"])
                worksheet.append([])
            workbook.save(template)

            result = write_brand_template(
                brand="법무법인 태하",
                template_path=template,
                output_path=output,
                run_date="2026-07-01",
                rows=[{"date": "", "media": "Naver", "category": "형", "keyword": "형사", "cost": "100"}],
            )

            self.assertEqual(result.written_rows, 0)
            self.assertEqual(result.touched_sheets, [])

    def test_taeha_category_matching_is_exact_so_hyeong_does_not_enter_military_sheet(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "output.xlsx"
            workbook = Workbook()
            workbook.remove(workbook.active)
            for sheet_name in ("형(1일)", "군형사(1일)"):
                worksheet = workbook.create_sheet(sheet_name)
                worksheet.append(["키워드", "총비용"])
                worksheet.append([])
            workbook.save(template)

            result = write_brand_template(
                brand="법무법인 태하",
                template_path=template,
                output_path=output,
                run_date="2026-07-01",
                rows=[{"date": "20260630", "media": "Naver", "category": "형", "keyword": "형사변호사", "cost": "100"}],
            )

            self.assertEqual(result.written_rows, 1)
            saved = load_workbook(output, data_only=False)
            self.assertEqual(saved["형(1일)"]["A3"].value, "형사변호사")
            self.assertIsNone(saved["군형사(1일)"]["A3"].value)

    def test_ohyun_category_matching_is_exact_so_criminal_does_not_enter_military_sheet(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "output.xlsx"
            workbook = Workbook()
            workbook.remove(workbook.active)
            for sheet_name in ("ohcrime(형사)1일", "군형사1일"):
                worksheet = workbook.create_sheet(sheet_name)
                worksheet.append(["키워드", "총비용"])
                worksheet.append([])
            workbook.save(template)

            result = write_brand_template(
                brand="법무법인 오현",
                template_path=template,
                output_path=output,
                run_date="2026-07-01",
                rows=[{"date": "20260630", "media": "Naver", "category": "형사", "campaign_type": "파워링크", "keyword": "형사변호사", "cost": "100"}],
            )

            self.assertEqual(result.written_rows, 1)
            saved = load_workbook(output, data_only=False)
            self.assertEqual(saved["ohcrime(형사)1일"]["A3"].value, "형사변호사")
            self.assertIsNone(saved["군형사1일"]["A3"].value)

    def test_template_writer_dedupes_rows_within_each_sheet(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "output.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "ohcrime(형사)1일"
            worksheet.append(["키워드", "노출수", "클릭수", "클릭률", "클릭비용", "총비용"])
            worksheet.append([])
            workbook.save(template)

            duplicated = {
                "date": "20260630",
                "media": "Naver",
                "category": "형사",
                "campaign_type": "파워링크",
                "device": "모바일",
                "keyword": "형사변호사",
                "impressions": "100",
                "clicks": "4",
                "ctr": "0.04",
                "cpc": "2500",
                "cost": "10000",
            }

            result = write_brand_template(
                brand="법무법인 오현",
                template_path=template,
                output_path=output,
                run_date="2026-07-01",
                rows=[duplicated, dict(duplicated)],
            )

            self.assertEqual(result.written_rows, 1)
            saved = load_workbook(output, data_only=False)
            self.assertEqual(saved["ohcrime(형사)1일"]["A3"].value, "형사변호사")
            self.assertIsNone(saved["ohcrime(형사)1일"]["A4"].value)

    def test_ohyun_allows_report_date_as_run_date_for_manual_upload(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "output.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "ohcrime(형사)1일"
            worksheet.append(["키워드", "클릭수", "노출수", "클릭률", "총비용"])
            worksheet.append([])
            workbook.save(template)

            result = write_brand_template(
                brand="법무법인 오현",
                template_path=template,
                output_path=output,
                run_date="2026-07-01",
                rows=[
                    {
                        "date": "20260701",
                        "media": "Naver",
                        "category": "형사",
                        "campaign_type": "파워링크",
                        "keyword": "형사변호사",
                        "clicks": "4",
                        "impressions": "100",
                        "ctr": "0",
                        "cost": "10000",
                    }
                ],
            )

            self.assertEqual(result.written_rows, 1)
            saved = load_workbook(output, data_only=False)
            self.assertEqual(saved["ohcrime(형사)1일"]["A3"].value, "형사변호사")
            self.assertEqual(saved["ohcrime(형사)1일"]["D3"].value, 0.04)

    def test_upload_rules_can_classify_by_url(self):
        with TemporaryDirectory() as temp_dir:
            rules = Path(temp_dir) / "rules.csv"
            rules.write_text(
                "브랜드,순위,규칙,매칭값,카테고리,신뢰도,사용,메모\n"
                "법무법인 오현,1,지정 URL,ohcrime.com,형사,1,O,\n",
                encoding="utf-8-sig",
            )
            raw = Path(temp_dir) / "ohyun_raw.csv"
            raw.write_text(
                '"일일보고서_0920(2026.06.30.~2026.06.30.),1826631"\n'
                "일별,URL,캠페인유형,PC/모바일 매체,키워드,노출수,클릭수,클릭률(%),평균 CPC,총비용,평균노출순위\n"
                "2026.06.30.,http://www.ohcrime.com,파워링크,모바일,형사변호사,100,1,1,5000,5000,2.1\n",
                encoding="utf-8-sig",
            )

            rows = build_upload_rows_from_raw(brand="법무법인 오현", media="Naver", input_path=raw, rules_path=rules)

            self.assertEqual(rows[0]["category"], "형사")

    def test_cpc_is_calculated_from_cost_and_clicks(self):
        with TemporaryDirectory() as temp_dir:
            rules = Path(temp_dir) / "rules.csv"
            rules.write_text(
                "브랜드,순위,규칙,매칭값,카테고리,신뢰도,사용,메모\n"
                "법무법인 오현,1,지정 URL,ohcrime.com,형사,1,O,\n",
                encoding="utf-8-sig",
            )
            raw = Path(temp_dir) / "ohyun_raw.csv"
            raw.write_text(
                '"일일보고서_0920(2026.06.30.~2026.06.30.),1826631"\n'
                "일별,URL,캠페인유형,PC/모바일 매체,키워드,노출수,클릭수,클릭률(%),평균 CPC,총비용,평균노출순위\n"
                "2026.06.30.,http://www.ohcrime.com,파워링크,모바일,형사변호사,100,4,1,999,10000,2.1\n",
                encoding="utf-8-sig",
            )

            rows = build_upload_rows_from_raw(brand="법무법인 오현", media="Naver", input_path=raw, rules_path=rules)

            self.assertEqual(rows[0]["cpc"], 2500)
            self.assertEqual(rows[0]["ctr"], 0.04)

    def test_ohyun_powercontents_can_use_campaign_as_keyword_and_category_rule(self):
        with TemporaryDirectory() as temp_dir:
            rules = Path(temp_dir) / "rules.csv"
            rules.write_text(
                "브랜드,매체,순위,규칙,매칭값,카테고리,신뢰도,사용,메모\n"
                "법무법인 오현,Naver,2,캠페인명,파워컨텐츠_교통,교통사고,1,O,\n",
                encoding="utf-8-sig",
            )
            raw = Path(temp_dir) / "ohyun_raw_20260630_20260630.csv"
            raw.write_text(
                '"일일보고서_0920(2026.06.30.~2026.06.30.),1826631"\n'
                "일별,캠페인명,캠페인유형,PC/모바일 매체,키워드,노출수,클릭수,클릭률(%),평균 CPC,총비용,평균노출순위\n"
                "2026.06.30.,파워컨텐츠_교통,파워컨텐츠,모바일,-,30,6,20,2786,16714,1.0\n",
                encoding="utf-8-sig",
            )

            rows = build_upload_rows_from_raw(brand="법무법인 오현", media="Naver", input_path=raw, rules_path=rules)

            self.assertEqual(rows[0]["category"], "교통사고")
            self.assertEqual(rows[0]["keyword"], "파워컨텐츠_교통")

    def test_ohyun_normal_and_spread_keyword_files_have_separate_roles(self):
        with TemporaryDirectory() as temp_dir:
            rules = Path(temp_dir) / "rules.csv"
            rules.write_text(
                "브랜드,매체,순위,규칙,매칭값,카테고리,신뢰도,사용,메모\n"
                "법무법인 오현,Naver,1,지정 URL,ohcrime.com,형사,1,O,\n"
                "법무법인 오현,Naver,2,캠페인명,파워컨텐츠_교통,교통사고,1,O,\n",
                encoding="utf-8-sig",
            )
            spread = Path(temp_dir) / "오현 키워드 소진액(스프레드),2030193.csv"
            spread.write_text(
                '"일일보고서_0920(2026.06.30.~2026.06.30.),1826631"\n'
                "일별,캠페인,광고그룹,URL,캠페인유형,PC/모바일 매체,키워드,노출수,클릭수,총비용,평균노출순위\n"
                "2026.06.30.,검색캠페인,그룹,http://www.ohcrime.com,파워링크,모바일,형사변호사,10,1,1000,1.0\n"
                "2026.06.30.,파워컨텐츠_교통,그룹,http://x,파워컨텐츠,모바일,보복운전처벌,10,1,2000,1.0\n",
                encoding="utf-8-sig",
            )
            normal = Path(temp_dir) / "오현 키워드 소진액,2030193.csv"
            normal.write_text(
                '"일일보고서_0920(2026.06.30.~2026.06.30.),1826631"\n'
                "일별,URL,캠페인유형,PC/모바일 매체,키워드,노출수,클릭수,클릭률(%),평균 CPC,총비용,평균노출순위\n"
                "2026.06.30.,http://www.ohcrime.com,파워링크,모바일,형사변호사,10,1,10,1000,1000,1.0\n"
                "2026.06.30.,http://x,파워컨텐츠,모바일,보복운전처벌,10,1,10,2000,2000,1.0\n",
                encoding="utf-8-sig",
            )

            spread_rows = build_upload_rows_from_raw(brand="법무법인 오현", media="Naver", input_path=spread, rules_path=rules)
            normal_rows = build_upload_rows_from_raw(brand="법무법인 오현", media="Naver", input_path=normal, rules_path=rules)

            self.assertEqual([(row["campaign_type"], row["category"], row["cost"]) for row in spread_rows], [("파워컨텐츠", "교통사고", 2000.0)])
            self.assertEqual([(row["campaign_type"], row["category"], row["cost"]) for row in normal_rows], [("파워링크", "형사", 1000.0)])

    def test_ohyun_powercontents_campaign_rule_can_match_powercontents_keyword(self):
        with TemporaryDirectory() as temp_dir:
            rules = Path(temp_dir) / "rules.csv"
            rules.write_text(
                "브랜드,매체,순위,규칙,매칭값,카테고리,신뢰도,사용,메모\n"
                "법무법인 오현,Naver,2,캠페인명,파워컨텐츠_교통,교통사고,1,O,\n",
                encoding="utf-8-sig",
            )
            raw = Path(temp_dir) / "ohyun_raw_20260630_20260630.csv"
            raw.write_text(
                '"일일보고서_0920(2026.06.30.~2026.06.30.),1826631"\n'
                "일별,캠페인명,캠페인유형,PC/모바일 매체,키워드,노출수,클릭수,클릭률(%),평균 CPC,총비용,평균노출순위\n"
                "2026.06.30.,오현_컨텐츠,파워컨텐츠,모바일,파워컨텐츠_교통,30,6,20,2786,16714,1.0\n",
                encoding="utf-8-sig",
            )

            rows = build_upload_rows_from_raw(brand="법무법인 오현", media="Naver", input_path=raw, rules_path=rules)

            self.assertEqual(rows[0]["category"], "교통사고")
            self.assertEqual(rows[0]["keyword"], "파워컨텐츠_교통")

    def test_upload_rules_can_classify_by_campaign_name(self):
        with TemporaryDirectory() as temp_dir:
            rules = Path(temp_dir) / "rules.csv"
            rules.write_text(
                "브랜드,순위,규칙,매칭값,카테고리,신뢰도,사용,메모\n"
                "법무법인 태하,2,캠페인명,성범죄,성범죄,1,O,\n",
                encoding="utf-8-sig",
            )
            raw = Path(temp_dir) / "google_sa_법무법인태하_01_형사_raw_20260630_20260630.csv"
            raw.write_text(
                "일간 키워드 보고\n"
                "2026년 6월 30일 - 2026년 6월 30일\n"
                "일\t캠페인\t캠페인 유형\t기기\t검색 키워드\t노출수\t클릭수\t클릭률(CTR)\t통화 코드\t평균 비용\t비용\t전환\t모든 전환당 비용\t전환율\n"
                "2026-06-30\t[SA] 서울_확장_성범죄_0317\t검색\t휴대전화\t성범죄변호사\t10\t1\t10.00%\tKRW\t1000\t1000\t0\t0\t0.00%\n",
                encoding="utf-16",
            )

            rows = build_upload_rows_from_raw(brand="법무법인 태하", media="Google SA", input_path=raw, rules_path=rules)

            self.assertEqual(rows[0]["category"], "성범죄")

    def test_google_campaign_rules_can_merge_to_same_sheet_category(self):
        with TemporaryDirectory() as temp_dir:
            rules = Path(temp_dir) / "rules.csv"
            rules.write_text(
                "브랜드,순위,규칙,매칭값,카테고리,신뢰도,사용,메모\n"
                "법무법인 태하,2,캠페인명,재산범죄,형,1,O,\n"
                "법무법인 태하,2,캠페인명,조세,형,1,O,\n",
                encoding="utf-8-sig",
            )
            raw = Path(temp_dir) / "google_sa_법무법인태하_01_형사_raw_20260630_20260630.csv"
            raw.write_text(
                "일간 키워드 보고\n"
                "2026년 6월 30일 - 2026년 6월 30일\n"
                "일\t캠페인\t캠페인 유형\t기기\t검색 키워드\t노출수\t클릭수\t클릭률(CTR)\t통화 코드\t평균 비용\t비용\t전환\t모든 전환당 비용\t전환율\n"
                "2026-06-30\t[SA] 서울_확장_재산범죄_0317\t검색\t휴대전화\t사기변호사\t10\t1\t10.00%\tKRW\t1000\t1000\t0\t0\t0.00%\n"
                "2026-06-30\t[SA] 조세_확장_0317\t검색\t휴대전화\t조세변호사\t10\t1\t10.00%\tKRW\t2000\t2000\t0\t0\t0.00%\n",
                encoding="utf-16",
            )

            rows = build_upload_rows_from_raw(brand="법무법인 태하", media="Google SA", input_path=raw, rules_path=rules)

            self.assertEqual([row["category"] for row in rows], ["형", "형"])

    def test_campaign_rules_can_be_limited_by_media(self):
        with TemporaryDirectory() as temp_dir:
            rules = Path(temp_dir) / "rules.csv"
            rules.write_text(
                "브랜드,매체,순위,규칙,매칭값,카테고리,신뢰도,사용,메모\n"
                "법무법인 태하,Naver,2,캠페인명,조세,조세,1,O,\n"
                "법무법인 태하,Google SA,2,캠페인명,조세,형,1,O,\n",
                encoding="utf-8-sig",
            )
            naver_raw = Path(temp_dir) / "naver_thlaw_04_raw_20260630_20260630.csv"
            naver_raw.write_text(
                '"일일보고서_0920(2026.06.30.~2026.06.30.),1826631"\n'
                "캠페인명,캠페인유형,PC/모바일 매체,키워드,노출수,클릭수,클릭률(%),평균 CPC,총비용,총 전환수,총 전환율(%),총 전환당비용(원),평균노출순위\n"
                "조세_확장,파워링크,모바일,조세변호사,100,1,1,5000,5000,0,0,0,2.1\n",
                encoding="utf-8-sig",
            )
            google_raw = Path(temp_dir) / "google_sa_법무법인태하_01_형사_raw_20260630_20260630.csv"
            google_raw.write_text(
                "일간 키워드 보고\n"
                "2026년 6월 30일 - 2026년 6월 30일\n"
                "일\t캠페인\t캠페인 유형\t기기\t검색 키워드\t노출수\t클릭수\t클릭률(CTR)\t통화 코드\t평균 비용\t비용\t전환\t모든 전환당 비용\t전환율\n"
                "2026-06-30\t[SA] 조세_확장_0317\t검색\t휴대전화\t조세변호사\t10\t1\t10.00%\tKRW\t2000\t2000\t0\t0\t0.00%\n",
                encoding="utf-16",
            )

            naver_rows = build_upload_rows_from_raw(brand="법무법인 태하", media="Naver", input_path=naver_raw, rules_path=rules)
            google_rows = build_upload_rows_from_raw(brand="법무법인 태하", media="Google SA", input_path=google_raw, rules_path=rules)

            self.assertEqual(naver_rows[0]["category"], "조세")
            self.assertEqual(google_rows[0]["category"], "형")

    def test_upload_rows_sort_by_cost_descending(self):
        rows = sort_upload_rows([
            {"keyword": "low", "cost": "100"},
            {"keyword": "high", "cost": "300"},
            {"keyword": "mid", "cost": "200"},
        ])

        self.assertEqual([row["keyword"] for row in rows], ["high", "mid", "low"])


if __name__ == "__main__":
    unittest.main()
