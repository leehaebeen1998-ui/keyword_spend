"""openpyxl로 최종 Excel 파일을 생성한다."""
from __future__ import annotations

import os
import subprocess
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from .aggregations import NUMBER_COLS_KR, RATE_COLS_KR


# ---------------------------------------------------------------------------
# 헤더 색상 (시트명 → hex)
# ---------------------------------------------------------------------------
_HEADER_COLORS: dict[str, str] = {
    "00_요약":              "1F4E79",   # 진파랑
    "SA 통합":              "2E75B6",   # 파랑
    "DA 통합":              "375623",   # 초록
    "90_Failed_Rows":       "C00000",   # 빨강
    "91_Unmapped_Columns":  "7030A0",   # 보라
    "99_Index_Log":         "595959",   # 회색
}
# SA 매체별 개별 시트 — 파랑 계열
_SA_MEDIA_COLOR = "4472C4"
# DA 매체별 개별 시트 — 초록 계열
_DA_MEDIA_COLOR = "548235"
_DEFAULT_HEADER_COLOR = "404040"

# AutoFilter를 적용할 시트 키워드 (시트명에 포함되면 적용)
_AUTOFILTER_KEYWORDS = ("SA 통합", "DA 통합", "SA", "DA", "Naver", "Google", "Kakao", "Meta")

# 시트 고정 순서
_SHEET_ORDER = [
    "00_요약",
    "SA 통합",
    "DA 통합",
    "90_Failed_Rows",
    "91_Unmapped_Columns",
    "99_Index_Log",
]


def build_workbook(
    sheet_data: dict[str, list[dict[str, Any]]],
    output_path: str | Path,
) -> Path:
    """sheet_data를 xlsx 파일로 저장한다."""
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl이 필요합니다. pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)

    # 고정 순서 + 매체별 개별 시트 (SA 통합 바로 뒤, DA 통합 바로 뒤)
    sa_media_sheets = [k for k in sheet_data
                       if k not in _SHEET_ORDER
                       and k not in ("90_Failed_Rows", "91_Unmapped_Columns", "99_Index_Log")
                       and _is_sa_sheet(k, sheet_data)]
    da_media_sheets = [k for k in sheet_data
                       if k not in _SHEET_ORDER
                       and k not in sa_media_sheets
                       and k not in ("90_Failed_Rows", "91_Unmapped_Columns", "99_Index_Log")]

    ordered: list[str] = []
    for name in _SHEET_ORDER:
        if name in sheet_data:
            ordered.append(name)
        if name == "SA 통합":
            ordered.extend(sorted(sa_media_sheets))
        elif name == "DA 통합":
            ordered.extend(sorted(da_media_sheets))

    # 나머지 (혹시 누락된 것)
    for name in sheet_data:
        if name not in ordered:
            ordered.append(name)

    for sheet_name in ordered:
        rows = sheet_data[sheet_name]
        _write_sheet(wb, sheet_name, rows)

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(out_path))
    except OSError:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)
        _move_with_powershell(Path(tmp_path), out_path)

    return out_path


def _is_sa_sheet(sheet_name: str, sheet_data: dict) -> bool:
    """시트가 SA 매체 데이터인지 판별.

    aggregations.py가 삽입한 __sheet_type__ 마커를 우선 참조.
    없으면 '소재' 컬럼 유무로 추정 (하위 호환).
    """
    rows = sheet_data.get(sheet_name, [])
    if not rows:
        return False
    first = rows[0]
    sheet_type = first.get("__sheet_type__")
    if sheet_type:
        return sheet_type == "SA"
    return "소재" not in first


def _should_autofilter(sheet_name: str) -> bool:
    """AutoFilter를 적용해야 하는 시트인지 판별."""
    skip = {"00_요약", "90_Failed_Rows", "91_Unmapped_Columns", "99_Index_Log"}
    return sheet_name not in skip


def _header_color(sheet_name: str, sheet_data: dict) -> str:
    if sheet_name in _HEADER_COLORS:
        return _HEADER_COLORS[sheet_name]
    if _is_sa_sheet(sheet_name, sheet_data):
        return _SA_MEDIA_COLOR
    return _DA_MEDIA_COLOR


def _write_sheet(
    wb: "Workbook",
    sheet_name: str,
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(title=sheet_name[:31])

    if not rows:
        ws.append(["(데이터 없음)"])
        return

    # 컬럼 목록
    _SKIP_KEYS = frozenset({"__dims_set__", "__sheet_type__"})
    all_keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen and key not in _SKIP_KEYS:
                all_keys.append(key)
                seen.add(key)

    # 헤더 색상
    color = _HEADER_COLORS.get(sheet_name, _DEFAULT_HEADER_COLOR)
    # 매체별 개별 시트는 별도 색상
    if sheet_name not in _HEADER_COLORS:
        color = _SA_MEDIA_COLOR if _is_sa_sheet(sheet_name, {sheet_name: rows}) else _DA_MEDIA_COLOR

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor=color)
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(all_keys)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append([_format_cell(k, row.get(k, "")) for k in all_keys])

    # 열 너비 자동 조정
    for col_idx, key in enumerate(all_keys, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(key)),
            *(len(str(_format_cell(key, row.get(key, "")))) for row in rows[:50]),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # AutoFilter — 매체 컬럼 드롭다운이 슬라이서 역할
    if _should_autofilter(sheet_name) and len(rows) > 0:
        last_col = get_column_letter(len(all_keys))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


def _format_cell(key: str, value: Any) -> Any:
    if value in (None, ""):
        return ""
    if key in NUMBER_COLS_KR:
        try:
            f = float(str(value).replace(",", ""))
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            return value
    if key in RATE_COLS_KR:
        try:
            return round(float(str(value).replace(",", "")), 4)
        except (ValueError, TypeError):
            return value
    return value


def _move_with_powershell(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    cmd = f'Move-Item -Path "{src}" -Destination "{dst}" -Force'
    result = subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", cmd],
        capture_output=True, text=True, check=False,
    )
    if result.returncode != 0:
        raise OSError(result.stderr or "PowerShell 파일 이동 실패")
