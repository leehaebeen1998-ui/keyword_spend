"""매체별 Raw Excel 빌더.

집계 없이 표준화된 행 원본을 그대로 출력.

고정 컬럼: 일자 / 광고유형 / 캠페인 / 그룹 / 기기 / 키워드 / 노출수 / 클릭수 / 비용 / 전환수
동적 컬럼: 전환 세부 항목(전화전환수 / 카카오전환수 / 문의전환수 등) — 해당 시트에 값이 있을 때만 추가
고정 컬럼: 카테고리

시트 구성: 매체별 1개 (Naver, Google SA, Kakao SA ...)
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 고정 컬럼 정의
# ---------------------------------------------------------------------------

# 앞쪽 고정 컬럼 (전환 세부 앞)
_FIXED_FRONT: list[tuple[str, str]] = [
    ("date",             "일자"),
    ("campaign_type",    "광고유형"),
    ("campaign_name",    "캠페인"),
    ("group_name",       "그룹"),
    ("device",           "기기"),
    ("keyword_name",     "키워드"),
    ("impressions",      "노출수"),
    ("clicks",           "클릭수"),
    ("cost",             "비용"),
    ("conversion_count", "전환수"),
]

# 전환 세부 항목 (값이 있는 컬럼만 동적으로 포함, 순서 고정)
CONVERSION_DETAIL_COLS: list[tuple[str, str]] = [
    ("phone_conversion_count",            "전화전환수"),
    ("kakao_conversion_count",            "카카오전환수"),
    ("general_inquiry_conversion_count",  "문의전환수"),
    ("channel_talk_conversion_count",     "채팅전환수"),
    ("db_conversion_count",               "DB전환수"),
    ("purchase_conversion_count",         "구매전환수"),
    ("purchase_conversion_revenue",       "구매전환액"),
]

# 뒤쪽 고정 컬럼
_FIXED_BACK: list[tuple[str, str]] = [
    ("category", "카테고리"),
]

_INT_COLS = {
    "impressions", "clicks", "cost", "conversion_count",
    "phone_conversion_count", "kakao_conversion_count",
    "general_inquiry_conversion_count", "channel_talk_conversion_count",
    "db_conversion_count", "purchase_conversion_count",
}
_FLOAT_COLS = {"purchase_conversion_revenue"}

# 시트 헤더 색상
_HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_BODY_FONT    = Font(size=10)
_NUM_ALIGN    = Alignment(horizontal="right")
_CENTER_ALIGN = Alignment(horizontal="center")


def build_raw_excel(
    standard_rows: list[dict[str, Any]],
    output_path: str | Path,
) -> Path:
    """표준화 행 → 매체별 Raw Excel 생성.

    Args:
        standard_rows: run_report_pipeline의 cleaned_rows
        output_path:   저장할 .xlsx 경로

    Returns:
        실제 저장된 Path
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 매체별 그룹핑 (MEDIA_ORDER 순서 유지)
    by_media: dict[str, list[dict]] = defaultdict(list)
    for row in standard_rows:
        media = str(row.get("media") or "Unknown")
        by_media[media].append(row)

    # MEDIA_ORDER 기준 정렬 (없으면 가나다순)
    try:
        from index_classifier.excel_builder.aggregations import SA_MEDIA, DA_MEDIA
        order = list(SA_MEDIA) + list(DA_MEDIA)
    except ImportError:
        order = []

    sorted_media = sorted(
        by_media.keys(),
        key=lambda m: (order.index(m) if m in order else 999, m),
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    for media in sorted_media:
        rows = by_media[media]
        sheet_name = _safe_name(media)
        ws = wb.create_sheet(title=sheet_name)
        # 해당 시트에서 값이 있는 전환 세부 컬럼만 선택
        active_detail = _active_conversion_cols(rows)
        _write_sheet(ws, rows, active_detail)

    if not wb.sheetnames:
        ws = wb.create_sheet("데이터 없음")

    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# 내부 유틸
# ---------------------------------------------------------------------------

def _safe_name(name: str) -> str:
    for ch in r'\/:*?[]':
        name = name.replace(ch, "")
    return name[:31]


def _active_conversion_cols(rows: list[dict]) -> list[tuple[str, str]]:
    """전환 세부 컬럼 중 해당 시트에 실제 값(>0)이 있는 것만 반환."""
    active = []
    for field, kr in CONVERSION_DETAIL_COLS:
        has_value = any(
            row.get(field) not in (None, "", 0, 0.0)
            for row in rows
        )
        if has_value:
            active.append((field, kr))
    return active


def _write_sheet(ws, rows: list[dict], detail_cols: list[tuple[str, str]]) -> None:
    # 최종 컬럼 = 앞고정 + 동적전환 + 뒤고정
    all_cols: list[tuple[str, str]] = _FIXED_FRONT + detail_cols + _FIXED_BACK

    # ── 헤더 ──
    for col_idx, (_, kr) in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=kr)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER_ALIGN

    # ── 데이터 ──
    for row_idx, row in enumerate(rows, 2):
        for col_idx, (field, _) in enumerate(all_cols, 1):
            val = row.get(field)
            if field in _INT_COLS:
                try:
                    val = int(float(val)) if val not in (None, "", 0) else 0
                except (ValueError, TypeError):
                    val = 0
            elif field in _FLOAT_COLS:
                try:
                    val = round(float(val), 0) if val not in (None, "", 0) else 0
                except (ValueError, TypeError):
                    val = 0
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _BODY_FONT
            if field in _INT_COLS or field in _FLOAT_COLS:
                cell.alignment = _NUM_ALIGN
                cell.number_format = "#,##0"

    # ── AutoFilter + 열 너비 ──
    if rows:
        last_col = get_column_letter(len(all_cols))
        ws.auto_filter.ref = f"A1:{last_col}1"

    col_widths = {
        "일자": 13, "광고유형": 12, "캠페인": 40, "그룹": 30,
        "기기": 10, "키워드": 25, "노출수": 10, "클릭수": 9,
        "비용": 12, "전환수": 9, "카테고리": 14,
        "전화전환수": 11, "카카오전환수": 12, "문의전환수": 11,
        "채팅전환수": 11, "DB전환수": 10, "구매전환수": 11, "구매전환액": 12,
    }
    for col_idx, (_, kr) in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(kr, 14)

    # 1행 고정
    ws.freeze_panes = "A2"
