from __future__ import annotations

import base64
import csv
import io
import os
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import Any


def _extract_zip_first_file(path: Path) -> Path:
    """zip 내 첫 번째 CSV/xlsx 파일을 임시 경로에 추출 후 반환."""
    with zipfile.ZipFile(path, "r") as zf:
        names = zf.namelist()
        target = next((n for n in names if n.lower().endswith(".csv")), None)
        if target is None:
            target = next((n for n in names if n.lower().endswith(".xlsx")), None)
        if target is None:
            raise ValueError(f"zip 내 CSV/xlsx 없음: {path}")
        data = zf.read(target)
    suffix = Path(target).suffix
    tmp = Path(tempfile.mktemp(suffix=suffix))
    tmp.write_bytes(data)
    return tmp


def read_raw_with_meta(
    path: str | Path,
    *,
    delimiter: str = ",",
) -> tuple[list[list[str]], int, str]:
    """원본 파일을 메타 행 포함 전체 rows로 읽고 헤더 위치를 반환.

    Returns:
        (all_rows, header_index, encoding)
        - all_rows: 파일 전체 행 (메타 행 포함)
        - header_index: 헤더 행 인덱스
        - encoding: 실제 사용한 인코딩
    """
    report_path = Path(path)

    if report_path.suffix.casefold() == ".zip":
        tmp = _extract_zip_first_file(report_path)
        try:
            return read_raw_with_meta(tmp, delimiter=delimiter)
        finally:
            tmp.unlink(missing_ok=True)

    encoding = _detect_encoding(report_path)
    # CSV는 탭/쉼표 자동 감지 (Google Ads 'Excel .csv'는 실제로 TSV)
    if delimiter == ",":
        delimiter = _detect_delimiter(report_path, encoding)
    try:
        with report_path.open("r", encoding=encoding, newline="") as file:
            rows = list(csv.reader(file, delimiter=delimiter))
    except UnicodeDecodeError:
        encoding = "cp949"
        with report_path.open("r", encoding=encoding, newline="") as file:
            rows = list(csv.reader(file, delimiter=delimiter))

    header_index = _detect_header_row(rows)
    return rows, header_index, encoding


def read_report_rows(path: str | Path, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
    report_path = Path(path)
    suffix = report_path.suffix.casefold()

    if suffix == ".zip":
        tmp = _extract_zip_first_file(report_path)
        try:
            return read_report_rows(tmp, sheet_name=sheet_name)
        finally:
            tmp.unlink(missing_ok=True)
    if suffix == ".csv":
        return _read_delimited(report_path, delimiter=",")
    if suffix == ".tsv":
        return _read_delimited(report_path, delimiter="\t")
    if suffix == ".xlsx":
        return _read_xlsx(report_path, sheet_name=sheet_name)

    raise ValueError(f"Unsupported report file type: {report_path.suffix}")


def write_csv_rows(
    path: str | Path,
    rows: list[dict[str, Any]],
    *,
    fieldnames: list[str],
    encoding: str = "utf-8-sig",
) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    file = io.StringIO(newline="")
    writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    write_text(output_path, file.getvalue(), encoding=encoding)


def write_text(path: str | Path, content: str, *, encoding: str = "utf-8") -> None:
    output_path = Path(path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with output_path.open("w", encoding=encoding, newline="") as file:
            file.write(content)
    except OSError:
        _write_text_with_powershell(output_path, content, encoding=encoding)


def _write_text_with_powershell(path: Path, content: str, *, encoding: str) -> None:
    command = (
        "$target = $env:INDEX_CLASSIFIER_TARGET_PATH; "
        "$directory = [System.IO.Path]::GetDirectoryName($target); "
        "if ($directory) { [System.IO.Directory]::CreateDirectory($directory) | Out-Null }; "
        "$base64 = [Console]::In.ReadToEnd(); "
        "$bytes = [Convert]::FromBase64String($base64); "
        "[System.IO.File]::WriteAllBytes($target, $bytes)"
    )
    encoded = base64.b64encode(content.encode(encoding)).decode("ascii")
    env = os.environ.copy()
    env["INDEX_CLASSIFIER_TARGET_PATH"] = os.fspath(path)
    result = subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", command],
        input=encoded,
        text=True,
        capture_output=True,
        check=False,
        env=env,
        encoding="utf-8",
        errors="replace",
    )
    if result.returncode != 0:
        message = (result.stderr or result.stdout or "PowerShell file write failed").strip()
        raise OSError(message)


def _detect_encoding(path: Path) -> str:
    """파일 BOM으로 인코딩 감지. UTF-16(Google Ads) 우선 처리."""
    try:
        raw = path.read_bytes()[:4]
    except OSError:
        return "utf-8-sig"
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16"
    if raw[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    return "utf-8-sig"


def _detect_delimiter(path: Path, encoding: str) -> str:
    """파일 앞부분 샘플로 탭/쉼표 수를 비교해 구분자 자동 감지.

    Google Ads 'Excel .csv' 내보내기는 실제로 탭 구분자(TSV)를 사용.
    """
    try:
        with path.open("r", encoding=encoding, newline="", errors="replace") as file:
            sample = file.read(8192)
        tab_count = sample.count("\t")
        comma_count = sample.count(",")
        return "\t" if tab_count > comma_count else ","
    except Exception:
        return ","


def _read_delimited(path: Path, *, delimiter: str) -> list[dict[str, Any]]:
    encoding = _detect_encoding(path)
    # CSV는 탭/쉼표 자동 감지 (Google Ads 'Excel .csv'는 실제로 TSV)
    if delimiter == ",":
        delimiter = _detect_delimiter(path, encoding)
    try:
        with path.open("r", encoding=encoding, newline="") as file:
            rows = list(csv.reader(file, delimiter=delimiter))
    except UnicodeDecodeError:
        encoding = "cp949"
        with path.open("r", encoding=encoding, newline="") as file:
            rows = list(csv.reader(file, delimiter=delimiter))

    if not rows:
        return []

    header_index = _detect_header_row(rows)
    headers = rows[header_index]
    results: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        item = {
            header: row[index] if index < len(row) else ""
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in item.values()):
            results.append(item)
    return results


def _detect_header_row(rows: list[list[str]]) -> int:
    known_headers = {
        # Naver
        "일별",
        "캠페인유형",
        "URL",
        "PC/모바일 매체",
        "캠페인",
        "광고그룹",
        "키워드",
        "노출수",
        "클릭수",
        "총비용",
        "총 전환수",
        # Google SA / DA
        "캠페인 유형",
        "검색 키워드",
        "기기",
        "비용",
        "전환",
        "일",
        "광고그룹",
        # Kakao SA
        "유형",
        "비즈채널명",
        "시작일",
        # Meta
        "광고 이름",
        "광고 세트 이름",
        "지출 금액 (KRW)",
    }

    for index, row in enumerate(rows[:20]):
        values = {str(value).strip() for value in row if str(value).strip()}
        if len(values & known_headers) >= 2:
            return index

    for index, row in enumerate(rows[:20]):
        non_empty = [value for value in row if str(value).strip()]
        if len(non_empty) > 2:
            return index

    return 0


def _read_xlsx(path: Path, *, sheet_name: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx report files.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = worksheet.iter_rows(values_only=True)

    # 헤더 행 감지 (메타 행이 있을 수 있음)
    all_rows = list(rows)
    if not all_rows:
        return []

    str_rows = [[str(v) if v is not None else "" for v in row] for row in all_rows]
    header_index = _detect_header_row(str_rows)

    headers = ["" if all_rows[header_index][i] is None else str(all_rows[header_index][i])
               for i in range(len(all_rows[header_index]))]
    results: list[dict[str, Any]] = []
    for row in all_rows[header_index + 1:]:
        item = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, "") for value in item.values()):
            results.append(item)
    return results
