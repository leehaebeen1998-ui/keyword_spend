from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .adapters.downloader_result_adapter import PipelineInput
from .classifier import ClassificationEngine
from .simple_rules import load_simple_rules_index
from .standardizer.media_column_mapping import load_default_mapping
from .standardizer.raw_standardizer import standardize_file

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


DAILY_REPORT_HEADERS: tuple[str, ...] = (
    "날짜",
    "브랜드",
    "계정",
    "매체",
    "카테고리",
    "캠페인",
    "그룹",
    "키워드",
    "URL",
    "노출",
    "클릭",
    "비용",
    "전환",
    "클릭률",
    "전환율",
    "전환당비용",
    "인덱스근거",
)


@dataclass(frozen=True)
class DailyReportResult:
    workbook_path: Path
    total_input_files: int
    total_rows: int
    failed_rows: int
    brand_sheets: list[str] = field(default_factory=list)


def build_daily_report(
    inputs: list[PipelineInput],
    rule_table_path: str | Path,
    output_path: str | Path,
    *,
    mapping_path: str | Path | None = None,
    decisions_path: str | Path | None = None,
    append: bool = True,
) -> DailyReportResult:
    """Raw 보고서를 브랜드별 데일리 보고 workbook으로 업데이트한다.

    브랜드별 시트에 카테고리 분류 결과와 주요 지표를 행 단위로 적재한다.
    클릭률/전환율/전환당비용은 Excel 수식으로 기록해, 데이터가 비어 있거나
    0인 경우에도 `IFERROR(...,0)`으로 회계/보고용 계산이 깨지지 않게 한다.
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl이 필요합니다. pip install openpyxl")

    if not inputs:
        raise ValueError("inputs가 비어 있습니다.")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    index = load_simple_rules_index(rule_table_path)
    engine = ClassificationEngine(index=index)
    mapping = load_default_mapping(mapping_path, decisions_path)

    rows_by_brand: dict[str, list[dict[str, Any]]] = {}
    failed_rows = 0
    total_rows = 0

    for pipeline_input in inputs:
        std_file = standardize_file(
            pipeline_input.file_path,
            media=pipeline_input.media,
            report_type=pipeline_input.report_type,
            account_name=pipeline_input.account_name,
            account_id=pipeline_input.account_id,
            brand_id=pipeline_input.brand_id,
            start_date=pipeline_input.start_date,
            end_date=pipeline_input.end_date,
            mapping=mapping,
            xlsx_sheet_name=pipeline_input.xlsx_sheet_name,
        )
        if std_file.error:
            raise RuntimeError(f"표준화 실패: {std_file.source_path} - {std_file.error}")

        for standard_row in std_file.standard_rows:
            result = engine.classify_row(standard_row)
            if not result.category:
                failed_rows += 1

            brand = _brand_name(standard_row, fallback=pipeline_input.brand_id or pipeline_input.account_name)
            daily_row = _build_daily_row(standard_row, result.to_dict(), brand)
            rows_by_brand.setdefault(brand, []).append(daily_row)
            total_rows += 1

    workbook = _load_or_create_workbook(output, append=append)
    brand_sheets: list[str] = []
    for brand, rows in sorted(rows_by_brand.items()):
        sheet_name = _safe_sheet_name(brand)
        brand_sheets.append(sheet_name)
        worksheet = _ensure_brand_sheet(workbook, sheet_name, append=append)
        _append_daily_rows(worksheet, rows)

    _write_summary_sheet(workbook, rows_by_brand, failed_rows)
    workbook.save(output)

    return DailyReportResult(
        workbook_path=output,
        total_input_files=len(inputs),
        total_rows=total_rows,
        failed_rows=failed_rows,
        brand_sheets=brand_sheets,
    )


def _load_or_create_workbook(path: Path, *, append: bool) -> "Workbook":
    if append and path.exists():
        return load_workbook(path)

    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def _ensure_brand_sheet(workbook: "Workbook", sheet_name: str, *, append: bool):
    if append and sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if worksheet.max_row == 0:
            _write_header(worksheet)
        return worksheet

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)
    _write_header(worksheet)
    return worksheet


def _write_header(worksheet) -> None:
    worksheet.append(list(DAILY_REPORT_HEADERS))
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"


def _append_daily_rows(worksheet, rows: list[dict[str, Any]]) -> None:
    rows = sorted(rows, key=lambda row: (
        str(row.get("카테고리") or ""),
        str(row.get("날짜") or ""),
        str(row.get("매체") or ""),
        str(row.get("캠페인") or ""),
    ))

    for row in rows:
        excel_row = worksheet.max_row + 1
        values = [row.get(header, "") for header in DAILY_REPORT_HEADERS]
        worksheet.append(values)

        worksheet.cell(excel_row, 14).value = f"=IFERROR(K{excel_row}/J{excel_row},0)"
        worksheet.cell(excel_row, 15).value = f"=IFERROR(M{excel_row}/K{excel_row},0)"
        worksheet.cell(excel_row, 16).value = f"=IFERROR(L{excel_row}/M{excel_row},0)"

        for col_idx in (10, 11, 12, 13, 16):
            worksheet.cell(excel_row, col_idx).number_format = '#,##0'
        for col_idx in (14, 15):
            worksheet.cell(excel_row, col_idx).number_format = "0.00%"

    _autosize_columns(worksheet)


def _write_summary_sheet(workbook: "Workbook", rows_by_brand: dict[str, list[dict[str, Any]]], failed_rows: int) -> None:
    if "00_요약" in workbook.sheetnames:
        del workbook["00_요약"]
    worksheet = workbook.create_sheet("00_요약", 0)
    worksheet.append(["브랜드", "카테고리", "행수", "노출", "클릭", "비용", "전환", "클릭률", "전환율", "전환당비용"])
    for cell in worksheet[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor="404040")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    summary: dict[tuple[str, str], dict[str, float]] = {}
    for brand, rows in rows_by_brand.items():
        for row in rows:
            key = (brand, str(row.get("카테고리") or "미분류"))
            metrics = summary.setdefault(key, {"rows": 0, "impressions": 0, "clicks": 0, "cost": 0, "conversions": 0})
            metrics["rows"] += 1
            metrics["impressions"] += _to_number(row.get("노출"))
            metrics["clicks"] += _to_number(row.get("클릭"))
            metrics["cost"] += _to_number(row.get("비용"))
            metrics["conversions"] += _to_number(row.get("전환"))

    for (brand, category), metrics in sorted(summary.items()):
        excel_row = worksheet.max_row + 1
        worksheet.append([
            brand,
            category,
            int(metrics["rows"]),
            metrics["impressions"],
            metrics["clicks"],
            metrics["cost"],
            metrics["conversions"],
            f"=IFERROR(E{excel_row}/D{excel_row},0)",
            f"=IFERROR(G{excel_row}/E{excel_row},0)",
            f"=IFERROR(F{excel_row}/G{excel_row},0)",
        ])
        for col_idx in (3, 4, 5, 6, 7, 10):
            worksheet.cell(excel_row, col_idx).number_format = "#,##0"
        for col_idx in (8, 9):
            worksheet.cell(excel_row, col_idx).number_format = "0.00%"

    worksheet.append([])
    worksheet.append(["미분류 행수", failed_rows])
    _autosize_columns(worksheet)


def _build_daily_row(standard_row: dict[str, Any], classification: dict[str, Any], brand: str) -> dict[str, Any]:
    category = classification.get("category") or "미분류"
    return {
        "날짜": standard_row.get("date") or "",
        "브랜드": brand,
        "계정": standard_row.get("account_name") or standard_row.get("account_id") or "",
        "매체": standard_row.get("media") or "",
        "카테고리": category,
        "캠페인": standard_row.get("campaign_name") or "",
        "그룹": standard_row.get("group_name") or "",
        "키워드": standard_row.get("keyword_name") or "",
        "URL": standard_row.get("url") or "",
        "노출": _to_number(standard_row.get("impressions")),
        "클릭": _to_number(standard_row.get("clicks")),
        "비용": _to_number(standard_row.get("cost")),
        "전환": _to_number(standard_row.get("conversion_count")),
        "인덱스근거": classification.get("matched_rule_id") or classification.get("source") or "",
    }


def _brand_name(row: dict[str, Any], *, fallback: str = "") -> str:
    value = row.get("brand_id") or fallback or row.get("account_name") or "default"
    return str(value).strip() or "default"


def _safe_sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", value).strip()
    return (cleaned or "default")[:31]


def _to_number(value: Any) -> float | int:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return value
    try:
        text = str(value).replace(",", "").replace("%", "").strip()
        number = float(text)
        return int(number) if number == int(number) else number
    except ValueError:
        return 0


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 42)
