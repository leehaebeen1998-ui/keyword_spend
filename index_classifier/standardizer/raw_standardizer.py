"""Raw 보고서 파일 → 표준 컬럼 변환 + 카테고리 annotate.

각 매체의 원본 컬럼명을 표준 컬럼명으로 변환한다.
annotate_file() 은 매체/채널 유형별 표준 출력 스키마로 저장한다.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from ..report_io import read_report_rows, read_raw_with_meta
from .media_column_mapping import (
    DIMENSION_COLUMNS,
    METRIC_COLUMNS,
    MediaColumnMapping,
    canonical_media,
    load_default_mapping,
)


# ---------------------------------------------------------------------------
# 출력 스키마 (std_field, 한국어_컬럼명)
# ---------------------------------------------------------------------------

#: SA 출력 컬럼 순서
_SA_OUTPUT_SCHEMA: list[tuple[str, str]] = [
    ("date",             "일자"),
    ("campaign_type",    "광고유형"),
    ("campaign_name",    "캠페인"),
    ("group_name",       "그룹"),
    ("device",           "기기"),
    ("keyword_name",     "키워드"),
    ("impressions",      "노출수"),
    ("clicks",           "클릭수"),
    ("cost",             "비용"),
    ("conversion_count", "전환수"),
]

#: DA 출력 컬럼 순서
_DA_OUTPUT_SCHEMA: list[tuple[str, str]] = [
    ("date",             "일자"),
    ("campaign_type",    "캠페인유형"),
    ("ad_type",          "광고유형"),
    ("campaign_name",    "캠페인"),
    ("group_name",       "그룹"),
    ("creative_name",    "소재"),
    ("device",           "기기"),
    ("keyword_name",     "키워드"),
    ("impressions",      "노출수"),
    ("clicks",           "클릭수"),
    ("cost",             "비용"),
    ("conversion_count", "전환수"),
    # 맞춤전환수 / 맞춤전환여부 는 _build_output_row 에서 파생 계산
]

#: DA 맞춤 전환 합산 대상 필드 (우선순위 순)
_DA_CUSTOM_CONV_FIELDS: list[str] = [
    "db_conversion_count",
    "general_inquiry_conversion_count",
    "phone_conversion_count",
    "purchase_conversion_count",
    "kakao_conversion_count",
    "channel_talk_conversion_count",
]


# ---------------------------------------------------------------------------
# 데이터 클래스
# ---------------------------------------------------------------------------

@dataclass
class UnmappedColumn:
    """매핑 없는 컬럼 정보."""
    media: str
    source_column: str
    sample_values: list[str] = field(default_factory=list)
    decision: str = "review"   # "map" | "ignore" | "review"
    suggested_target: str = ""


@dataclass
class StandardizedFile:
    """표준화 결과."""
    source_path: Path
    media: str
    report_type: str
    channel_type: str          # "SA" | "DA" | "GA4" | "unknown"
    account_name: str
    account_id: str
    brand_id: str
    start_date: str
    end_date: str
    standard_rows: list[dict[str, Any]]
    unmapped_columns: list[UnmappedColumn]
    row_count: int
    error: str | None = None


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def standardize_file(
    file_path: str | Path,
    *,
    media: str,
    report_type: str = "",
    account_name: str = "",
    account_id: str = "",
    brand_id: str = "",
    start_date: str = "",
    end_date: str = "",
    mapping: MediaColumnMapping | None = None,
    mapping_path: str | Path | None = None,
    decisions_path: str | Path | None = None,
    xlsx_sheet_name: str | None = None,
) -> StandardizedFile:
    """Raw 파일 하나를 표준 컬럼 행 목록으로 변환한다."""
    _mapping = mapping or load_default_mapping(mapping_path, decisions_path)
    source_path = Path(file_path)
    channel_type = _mapping.channel_type(media) or "unknown"

    try:
        raw_rows = read_report_rows(source_path, sheet_name=xlsx_sheet_name)
    except Exception as exc:
        return StandardizedFile(
            source_path=source_path,
            media=media,
            report_type=report_type or _infer_report_type(channel_type, media),
            channel_type=channel_type,
            account_name=account_name,
            account_id=account_id,
            brand_id=brand_id,
            start_date=start_date,
            end_date=end_date,
            standard_rows=[],
            unmapped_columns=[],
            row_count=0,
            error=str(exc),
        )

    if not raw_rows:
        return StandardizedFile(
            source_path=source_path,
            media=media,
            report_type=report_type or _infer_report_type(channel_type, media),
            channel_type=channel_type,
            account_name=account_name,
            account_id=account_id,
            brand_id=brand_id,
            start_date=start_date,
            end_date=end_date,
            standard_rows=[],
            unmapped_columns=[],
            row_count=0,
        )

    raw_columns = list(raw_rows[0].keys())
    unmapped = _detect_unmapped(_mapping, media, raw_columns, raw_rows[:5])

    if channel_type == "GA4":
        standard_rows = _standardize_ga4_rows(
            raw_rows,
            media=media,
            account_name=account_name,
            account_id=account_id,
            brand_id=brand_id,
            source_file=str(source_path),
            mapping=_mapping,
        )
    else:
        standard_rows = [
            _standardize_row(
                row,
                media=media,
                account_name=account_name,
                account_id=account_id,
                brand_id=brand_id,
                source_file=str(source_path),
                mapping=_mapping,
            )
            for row in raw_rows
        ]

    effective_report_type = report_type or _infer_report_type(channel_type, media)
    for row in standard_rows:
        row["report_type"] = effective_report_type
        row["channel_type"] = channel_type

    return StandardizedFile(
        source_path=source_path,
        media=media,
        report_type=effective_report_type,
        channel_type=channel_type,
        account_name=account_name,
        account_id=account_id,
        brand_id=brand_id,
        start_date=start_date,
        end_date=end_date,
        standard_rows=standard_rows,
        unmapped_columns=unmapped,
        row_count=len(standard_rows),
    )


def standardize_files(
    inputs: list[dict[str, Any]],
    *,
    mapping: MediaColumnMapping | None = None,
    mapping_path: str | Path | None = None,
    decisions_path: str | Path | None = None,
) -> list[StandardizedFile]:
    """여러 파일을 한꺼번에 표준화한다."""
    _mapping = mapping or load_default_mapping(mapping_path, decisions_path)
    results: list[StandardizedFile] = []
    for item in inputs:
        result = standardize_file(
            item["file_path"],
            media=item["media"],
            report_type=item.get("report_type", ""),
            account_name=item.get("account_name", ""),
            account_id=item.get("account_id", ""),
            brand_id=item.get("brand_id", ""),
            start_date=item.get("start_date", ""),
            end_date=item.get("end_date", ""),
            mapping=_mapping,
            xlsx_sheet_name=item.get("xlsx_sheet_name"),
        )
        results.append(result)
    return results


def annotate_file(
    file_path: str | Path,
    *,
    media: str,
    account_name: str = "",
    account_id: str = "",
    brand_id: str = "",
    mapping: MediaColumnMapping | None = None,
    mapping_path: str | Path | None = None,
    engine: Any,
    output_path: str | Path,
    category_column: str = "카테고리",
) -> None:
    """원본 파일을 표준 스키마로 변환하고 카테고리 컬럼을 붙여 저장.

    SA → _SA_OUTPUT_SCHEMA  (일자·광고유형·캠페인·그룹·소재·기기·키워드 + 지표 + 카테고리)
    DA → _DA_OUTPUT_SCHEMA  (위 + 캠페인유형·광고유형·맞춤전환수·맞춤전환여부 + 카테고리)
    """
    _mapping = mapping or load_default_mapping(mapping_path)
    channel_type = _mapping.channel_type(media) or "SA"
    schema = _DA_OUTPUT_SCHEMA if channel_type == "DA" else _SA_OUTPUT_SCHEMA

    src = Path(file_path)
    dst = Path(output_path)
    suffix = src.suffix.casefold()

    if suffix in (".csv", ".tsv"):
        delimiter = "," if suffix == ".csv" else "\t"
        _annotate_delimited(
            src, dst,
            delimiter=delimiter,
            media=media,
            account_name=account_name,
            account_id=account_id,
            brand_id=brand_id,
            mapping=_mapping,
            engine=engine,
            schema=schema,
            channel_type=channel_type,
            category_column=category_column,
        )
    elif suffix == ".xlsx":
        _annotate_xlsx(
            src, dst,
            media=media,
            account_name=account_name,
            account_id=account_id,
            brand_id=brand_id,
            mapping=_mapping,
            engine=engine,
            schema=schema,
            channel_type=channel_type,
            category_column=category_column,
        )
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {suffix}")


# ---------------------------------------------------------------------------
# 내부 — 표준화 로직
# ---------------------------------------------------------------------------

def _standardize_row(
    raw_row: dict[str, Any],
    *,
    media: str,
    account_name: str,
    account_id: str,
    brand_id: str,
    source_file: str,
    mapping: MediaColumnMapping,
) -> dict[str, Any]:
    """단일 raw 행을 표준 행으로 변환."""
    std: dict[str, Any] = {col: "" for col in DIMENSION_COLUMNS}
    std.update({col: 0 for col in METRIC_COLUMNS})

    std["media"] = media
    std["account_name"] = account_name
    std["account_id"] = account_id
    std["brand_id"] = brand_id
    std["source_file"] = source_file

    col_map = mapping.column_map.get(canonical_media(media), {})

    for raw_col, value in raw_row.items():
        target = col_map.get(raw_col)
        if not target:
            continue
        if target in METRIC_COLUMNS:
            if std.get(target) in (0, "", None):
                std[target] = _safe_number(value)
        else:
            if std.get(target) in ("", None):
                std[target] = value if value is not None else ""

    # campaign_type이 비어있으면 media 이름으로 채움
    if not std.get("campaign_type"):
        std["campaign_type"] = media

    return std


def _standardize_ga4_rows(
    raw_rows: list[dict[str, Any]],
    *,
    media: str,
    account_name: str,
    account_id: str,
    brand_id: str,
    source_file: str,
    mapping: MediaColumnMapping,
) -> list[dict[str, Any]]:
    """GA4 이벤트 피벗 처리."""
    col_map = mapping.column_map.get(media, {})
    evt_map = mapping.event_map.get(media, {})

    renamed: list[dict[str, Any]] = []
    for raw_row in raw_rows:
        row: dict[str, Any] = {}
        for raw_col, value in raw_row.items():
            target = col_map.get(raw_col, raw_col)
            row[target] = value
        row["media"] = media
        row["account_name"] = account_name
        row["account_id"] = account_id
        row["brand_id"] = brand_id
        row["source_file"] = source_file
        renamed.append(row)

    if not renamed or "conversion_event_name" not in renamed[0]:
        return renamed

    group_dim_cols = {"date", "media", "account_name", "account_id", "brand_id",
                      "creative_name", "keyword_name", "source_file"}

    from collections import defaultdict
    groups: dict[tuple, dict[str, Any]] = defaultdict(dict)
    unmapped_events: set[str] = set()

    for row in renamed:
        evt_name = str(row.get("conversion_event_name") or "")
        evt_value = _safe_number(row.get("conversion_event_value", 0))
        key_values = tuple(str(row.get(c, "")) for c in sorted(group_dim_cols))
        group = groups[key_values]
        for col in group_dim_cols:
            group.setdefault(col, row.get(col, ""))
        target_metric = evt_map.get(evt_name)
        if target_metric:
            group[target_metric] = group.get(target_metric, 0) + evt_value
        elif evt_name:
            unmapped_events.add(evt_name)

    result_rows: list[dict[str, Any]] = []
    for group in groups.values():
        std: dict[str, Any] = {col: "" for col in DIMENSION_COLUMNS}
        std.update({col: 0 for col in METRIC_COLUMNS})
        std.update(group)
        if unmapped_events:
            std["_ga4_unmapped_events"] = ", ".join(sorted(unmapped_events))
        result_rows.append(std)

    return result_rows


def _detect_unmapped(
    mapping: MediaColumnMapping,
    media: str,
    raw_columns: list[str],
    sample_rows: list[dict[str, Any]],
) -> list[UnmappedColumn]:
    col_map = mapping.column_map.get(media, {})
    result: list[UnmappedColumn] = []
    for col in raw_columns:
        if col in col_map:
            continue
        decision = mapping.decision_for(media, col) or "review"
        samples = [str(row.get(col, "")) for row in sample_rows
                   if row.get(col) not in (None, "")][:3]
        result.append(UnmappedColumn(
            media=media,
            source_column=col,
            sample_values=samples,
            decision=decision,
        ))
    return result


def _infer_report_type(channel_type: str, media: str) -> str:
    if channel_type == "SA":
        return "SA_KEYWORD"
    if channel_type == "DA":
        return "DA_CREATIVE"
    if channel_type == "GA4":
        return "GA4_EVENT"
    return "UNKNOWN"


# ---------------------------------------------------------------------------
# 내부 — annotate 출력 로직
# ---------------------------------------------------------------------------

def _build_output_row(
    std_row: dict[str, Any],
    schema: list[tuple[str, str]],
    category: str,
    channel_type: str,
    category_column: str,
) -> dict[str, Any]:
    """표준 행 → 출력 딕셔너리 (스키마 컬럼 + 카테고리)."""
    out: dict[str, Any] = {}
    for field_name, display in schema:
        val = std_row.get(field_name, "")
        # 숫자 필드 0 → 빈 문자열로 보이게 하지 않고 그대로 유지
        out[display] = val

    if channel_type == "DA":
        custom_conv = sum(
            _safe_number(std_row.get(f, 0)) for f in _DA_CUSTOM_CONV_FIELDS
        )
        out["맞춤전환수"] = int(custom_conv) if custom_conv == int(custom_conv) else custom_conv
        out["맞춤전환여부"] = "Y" if custom_conv > 0 else "N"

    out[category_column] = category
    return out


def _output_headers(
    schema: list[tuple[str, str]],
    channel_type: str,
    category_column: str,
) -> list[str]:
    """출력 헤더 목록 반환."""
    headers = [display for _, display in schema]
    if channel_type == "DA":
        headers += ["맞춤전환수", "맞춤전환여부"]
    headers.append(category_column)
    return headers


def _annotate_delimited(
    src: Path,
    dst: Path,
    *,
    delimiter: str,
    media: str,
    account_name: str,
    account_id: str,
    brand_id: str,
    mapping: MediaColumnMapping,
    engine: Any,
    schema: list[tuple[str, str]],
    channel_type: str,
    category_column: str,
) -> None:
    all_rows, header_index, _enc = read_raw_with_meta(src, delimiter=delimiter)
    if header_index >= len(all_rows):
        return

    raw_headers = all_rows[header_index]
    col_map = mapping.column_map.get(canonical_media(media), {})
    mapped = [h for h in raw_headers if h in col_map]
    unmapped = [h for h in raw_headers if h and h not in col_map]
    print(f"  [annotate-csv] {src.name} | media={media} | header_row={header_index}")
    print(f"    매핑됨({len(mapped)}): {mapped}")
    print(f"    미매핑({len(unmapped)}): {unmapped}")
    out_headers = _output_headers(schema, channel_type, category_column)

    output_rows: list[list[str]] = []

    # 헤더 행 이전 메타 행 (빈 행으로 보존)
    for i in range(header_index):
        output_rows.append([""] * len(out_headers))

    output_rows.append(out_headers)

    for row in all_rows[header_index + 1:]:
        # 완전히 빈 행
        if not any(str(v).strip() for v in row):
            output_rows.append([""] * len(out_headers))
            continue

        raw_row: dict[str, Any] = {
            raw_headers[j]: (row[j] if j < len(row) else "")
            for j in range(len(raw_headers))
        }
        std_row = _standardize_row(
            raw_row,
            media=media,
            account_name=account_name,
            account_id=account_id,
            brand_id=brand_id,
            source_file=str(src),
            mapping=mapping,
        )
        result = engine.classify_row(std_row)
        out_row = _build_output_row(std_row, schema, result.category or "", channel_type, category_column)
        output_rows.append([str(out_row.get(h, "")) for h in out_headers])

    dst.parent.mkdir(parents=True, exist_ok=True)
    with dst.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(output_rows)


def _annotate_xlsx(
    src: Path,
    dst: Path,
    *,
    media: str,
    account_name: str,
    account_id: str,
    brand_id: str,
    mapping: MediaColumnMapping,
    engine: Any,
    schema: list[tuple[str, str]],
    channel_type: str,
    category_column: str,
) -> None:
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl이 필요합니다.") from exc

    wb_src = load_workbook(src, read_only=True, data_only=True)
    ws_src = wb_src.active
    all_values = list(ws_src.iter_rows(values_only=True))
    if not all_values:
        return

    str_rows = [[str(v) if v is not None else "" for v in row] for row in all_values]
    from ..report_io import _detect_header_row
    header_index = _detect_header_row(str_rows)

    raw_headers = [str(v) if v is not None else "" for v in all_values[header_index]]
    col_map = mapping.column_map.get(canonical_media(media), {})
    mapped = [h for h in raw_headers if h in col_map]
    unmapped = [h for h in raw_headers if h and h not in col_map]
    print(f"  [annotate-xlsx] {src.name} | media={media} | header_row={header_index}")
    print(f"    매핑됨({len(mapped)}): {mapped}")
    print(f"    미매핑({len(unmapped)}): {unmapped}")
    out_headers = _output_headers(schema, channel_type, category_column)

    wb_dst = Workbook()
    ws_dst = wb_dst.active

    # 헤더 행 이전 메타 행 보존 (빈 셀)
    for _i in range(header_index):
        ws_dst.append([""] * len(out_headers))

    ws_dst.append(out_headers)

    for row_values in all_values[header_index + 1:]:
        if not any(v is not None and str(v).strip() for v in row_values):
            ws_dst.append([""] * len(out_headers))
            continue

        raw_row: dict[str, Any] = {
            raw_headers[j]: (
                str(row_values[j]) if j < len(row_values) and row_values[j] is not None else ""
            )
            for j in range(len(raw_headers))
        }
        std_row = _standardize_row(
            raw_row,
            media=media,
            account_name=account_name,
            account_id=account_id,
            brand_id=brand_id,
            source_file=str(src),
            mapping=mapping,
        )
        result = engine.classify_row(std_row)
        out_row = _build_output_row(std_row, schema, result.category or "", channel_type, category_column)
        ws_dst.append([out_row.get(h, "") for h in out_headers])

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb_dst.save(dst)


# ---------------------------------------------------------------------------
# 내부 유틸
# ---------------------------------------------------------------------------

def _safe_number(value: Any) -> float | int:
    """문자열 → 숫자 변환. 실패 시 0."""
    if isinstance(value, (int, float)):
        return value
    try:
        text = str(value).replace(",", "").strip()
        if "." in text:
            return float(text)
        return int(text)
    except (ValueError, TypeError):
        return 0
