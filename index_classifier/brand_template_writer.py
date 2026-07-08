from __future__ import annotations

import csv
import json
import os
import re
import subprocess
import tempfile
from dataclasses import dataclass, replace
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from .brand_upload import BrandUploadRule, DEFAULT_BRAND_RULES, SheetTarget, build_sheet_targets, parse_run_date


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "date": ("일자", "날짜", "일별", "date"),
    "media": ("매체", "media"),
    "category": ("카테고리", "분류", "category"),
    "campaign_type": ("캠페인유형", "campaign_type"),
    "device": ("PC/모바일", "device"),
    "keyword": ("키워드", "keyword", "keyword_name"),
    "impressions": ("노출수", "impressions"),
    "clicks": ("클릭수", "clicks"),
    "ctr": ("클릭률", "ctr"),
    "cpc": ("클릭비용", "cpc"),
    "cost": ("총비용", "cost"),
    "conversions": ("전환수", "conversions", "conversion_count"),
    "conversion_rate": ("전환율", "conversion_rate"),
    "cost_per_conversion": ("전환당비용", "cost_per_conversion", "cpa"),
    "rank": ("노출순위", "rank", "average_rank"),
}

NUMBER_FORMATS: dict[str, str] = {
    "conversions": "0;-0;-",
    "conversion_rate": "0.00%;-0.00%;-",
    "cost_per_conversion": "#,##0;-#,##0;-",
}


@dataclass(frozen=True)
class TemplateWriteResult:
    output_path: Path
    written_rows: int
    touched_sheets: list[str]
    skipped_rows: int
    expected_sheet_names: tuple[str, ...] = ()
    missing_sheets: tuple[str, ...] = ()
    available_row_dates: tuple[str, ...] = ()


def load_upload_rows(path: str | Path) -> list[dict[str, Any]]:
    csv_path = Path(path)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        return [dict(row) for row in csv.DictReader(file)]


def write_brand_template(
    *,
    brand: str,
    template_path: str | Path,
    output_path: str | Path,
    rows: list[dict[str, Any]],
    run_date: str | date | datetime | None = None,
    update_mode: str = "replace",
    category_mode: str = "category_sheets",
    rule: BrandUploadRule | None = None,
) -> TemplateWriteResult:
    rule = rule if rule is not None else _brand_rule_for_upload(brand=brand, rows=rows)
    template = Path(template_path)
    output = Path(output_path)
    suffix = template.suffix.lower()

    if suffix == ".xlsx":
        return write_xlsx_template(rule=rule, template_path=template, output_path=output, rows=rows, run_date=run_date, update_mode=update_mode, category_mode=category_mode)
    if suffix == ".xlsb":
        return write_xlsb_template_with_excel(rule=rule, template_path=template, output_path=output, rows=rows, run_date=run_date, update_mode=update_mode, category_mode=category_mode)
    raise ValueError(f"Unsupported template type: {template.suffix}")


def _brand_rule_for_upload(*, brand: str, rows: list[dict[str, Any]]) -> BrandUploadRule:
    if brand in DEFAULT_BRAND_RULES:
        return DEFAULT_BRAND_RULES[brand]
    categories = []
    seen: set[str] = set()
    for row in rows:
        category = str(_first(row, "category", "카테고리") or "").strip()
        if category and category not in seen:
            seen.add(category)
            categories.append(category)
    if not categories:
        categories = ["미분류"]
    return BrandUploadRule(
        brand=brand,
        mode="fixed_today_offset",
        categories=tuple(categories),
        today_offset=1,
        use_today_formula=True,
    )


def write_xlsx_template(
    *,
    rule: BrandUploadRule,
    template_path: str | Path,
    output_path: str | Path,
    rows: list[dict[str, Any]],
    run_date: str | date | datetime | None = None,
    update_mode: str = "replace",
    category_mode: str = "category_sheets",
) -> TemplateWriteResult:
    try:
        from copy import copy
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for .xlsx template writing.") from exc

    current = parse_run_date(run_date)
    workbook = load_workbook(template_path)
    targets = _single_sheet_targets(rows, rule, current) if category_mode == "single_sheet" else _build_sheet_targets_for_rows(rule, current, rows)
    touched: list[str] = []
    written = 0
    missing_sheets: list[str] = []

    for target in targets:
        sheet_name = _first_data_sheet_name(workbook) if target.sheet_name == "__SINGLE_SHEET__" else _resolve_workbook_sheet_name(workbook.sheetnames, target.sheet_name)
        if sheet_name is None:
            if target.sheet_name != "__SINGLE_SHEET__":
                missing_sheets.append(target.sheet_name)
            continue
        target_rows = _rows_for_target(rows, target)

        worksheet = workbook[sheet_name]
        header_row = _find_header_row_openpyxl(worksheet)
        if header_row is None:
            continue

        header_map = _header_map_openpyxl(worksheet, header_row)
        data_start = header_row + 2
        write_start = _append_start_row_openpyxl(worksheet, data_start, header_map) if update_mode == "append" else data_start
        template_row = data_start if worksheet.max_row >= data_start else header_row + 1
        if update_mode != "append":
            _clear_openpyxl_data(worksheet, data_start, header_map.values())
        if not target_rows:
            touched.append(sheet_name)
            continue

        for row_index, row in enumerate(target_rows, start=write_start):
            if row_index > worksheet.max_row:
                worksheet.append([])
            for field, col_idx in header_map.items():
                cell = worksheet.cell(row_index, col_idx)
                source_style = worksheet.cell(template_row, col_idx)
                if source_style.has_style:
                    cell._style = copy(source_style._style)
                    cell.number_format = source_style.number_format
                    cell.alignment = copy(source_style.alignment)
                    cell.font = copy(source_style.font)
                    cell.fill = copy(source_style.fill)
                    cell.border = copy(source_style.border)
                cell.value = _value_for_field(row, field)
                if field in NUMBER_FORMATS:
                    cell.number_format = NUMBER_FORMATS[field]
            written += 1
        touched.append(sheet_name)

    output = Path(output_path)
    _ensure_directory(output.parent)
    _save_openpyxl_workbook(workbook, output)
    available_row_dates = tuple(sorted({
        parsed.strftime("%Y-%m-%d")
        for parsed in (_parse_row_date(_first(row, "date", "날짜", "일자")) for row in rows)
        if parsed is not None
    }))
    expected_sheet_names = tuple(target.sheet_name for target in targets)
    return TemplateWriteResult(
        output_path=output,
        written_rows=written,
        touched_sheets=touched,
        skipped_rows=len(rows) - written,
        expected_sheet_names=expected_sheet_names,
        missing_sheets=tuple(missing_sheets),
        available_row_dates=available_row_dates,
    )


def write_xlsb_template_with_excel(
    *,
    rule: BrandUploadRule,
    template_path: str | Path,
    output_path: str | Path,
    rows: list[dict[str, Any]],
    run_date: str | date | datetime | None = None,
    update_mode: str = "replace",
    category_mode: str = "category_sheets",
) -> TemplateWriteResult:
    current = parse_run_date(run_date)
    sheet_targets = _single_sheet_targets(rows, rule, current) if category_mode == "single_sheet" else _build_sheet_targets_for_rows(rule, current, rows)
    targets = [
        {
            "sheet_name": target.sheet_name,
            "category": target.category,
            "report_date": target.report_date.strftime("%Y%m%d"),
            "media": target.media,
            "rows": [_prepare_write_row(row) for row in _rows_for_target(rows, target)],
        }
        for target in sheet_targets
    ]
    output = Path(output_path)
    _ensure_directory(output.parent)
    payload = {
        "template_path": os.fspath(Path(template_path)),
        "output_path": os.fspath(output),
        "targets": targets,
        "headers": HEADER_ALIASES,
        "update_mode": update_mode,
    }

    with tempfile.NamedTemporaryFile("w", suffix=".json", encoding="utf-8", delete=False) as file:
        json.dump(payload, file, ensure_ascii=False)
        payload_path = file.name

    try:
        try:
            result = subprocess.run(
                ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", _POWERSHELL_XLSB_WRITER],
                input=payload_path,
                text=True,
                capture_output=True,
                encoding="utf-8",
                errors="replace",
                check=False,
                timeout=300,
            )
        except subprocess.TimeoutExpired as exc:
            raise TimeoutError("Excel 템플릿 반영이 5분을 초과했습니다. Excel 백그라운드 프로세스 또는 열린 파일을 확인해 주세요.") from exc
        if result.returncode != 0:
            raise RuntimeError((result.stderr or result.stdout or "Excel COM writer failed").strip())
    finally:
        try:
            Path(payload_path).unlink()
        except OSError:
            pass

    touched = [target["sheet_name"] for target in targets]
    written = sum(len(target["rows"]) for target in targets)
    available_row_dates = tuple(sorted({
        parsed.strftime("%Y-%m-%d")
        for parsed in (_parse_row_date(_first(row, "date", "날짜", "일자")) for row in rows)
        if parsed is not None
    }))
    return TemplateWriteResult(
        output_path=output,
        written_rows=written,
        touched_sheets=touched,
        skipped_rows=len(rows) - written,
        expected_sheet_names=tuple(target["sheet_name"] for target in sheet_targets),
        available_row_dates=available_row_dates,
    )


def _rows_for_target(rows: Iterable[dict[str, Any]], target: SheetTarget) -> list[dict[str, Any]]:
    if target.sheet_name == "__SINGLE_SHEET__":
        result = [row for row in rows if _row_matches_single_sheet_target(row, target)]
        result = _merge_rows_for_sheet(result)
        result = [row for row in result if _has_positive_cost(row)]
        result.sort(key=lambda row: _to_number(_first(row, "cost", "총비용")), reverse=True)
        return result
    result = [row for row in rows if _row_matches_target(row, target)]
    result = _merge_rows_for_sheet(result)
    result = [row for row in result if _has_positive_cost(row)]
    result.sort(key=lambda row: _to_number(_first(row, "cost", "총비용")), reverse=True)
    return result


def _single_sheet_targets(rows: list[dict[str, Any]], rule: BrandUploadRule, current: date) -> list[SheetTarget]:
    dates = sorted({target.report_date for target in build_sheet_targets(rule, current)})
    if not dates:
        dates = sorted({_parse_row_date(_first(row, "date", "날짜", "일자")) for row in rows if _parse_row_date(_first(row, "date", "날짜", "일자"))})
    if not dates:
        dates = [current]
    return [
        SheetTarget(
            brand=rule.brand,
            category="__ALL_DATES__",
            report_date=dates[0],
            sheet_name="__SINGLE_SHEET__",
            media=",".join(target_date.strftime("%Y%m%d") for target_date in dates),
        )
    ]


def _row_matches_single_sheet_target(row: dict[str, Any], target: SheetTarget) -> bool:
    row_date = _parse_row_date(_first(row, "date", "날짜", "일자"))
    if not row_date:
        return False
    allowed = {value for value in str(target.media or "").split(",") if value}
    return row_date.strftime("%Y%m%d") in allowed if allowed else row_date == target.report_date


def _dedupe_rows_for_sheet(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, ...]] = set()
    result: list[dict[str, Any]] = []
    for row in rows:
        key = _sheet_dedupe_key(row)
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def _merge_rows_for_sheet(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in rows:
        key = _sheet_merge_key(row)
        if key not in merged:
            merged[key] = dict(row)
            continue
        _add_metrics(merged[key], row)
    return [_recalculate_metrics(row) for row in merged.values()]


def _sheet_merge_key(row: dict[str, Any]) -> tuple[str, ...]:
    return (
        _normalize(str(_first(row, "date", "날짜", "일자") or "")),
        _normalize(str(_first(row, "media", "매체") or "")),
        _normalize(str(_first(row, "category", "카테고리") or "")),
        _normalize(str(_first(row, "campaign_type", "캠페인유형") or "")),
        _normalize(str(_first(row, "device", "PC/모바일") or "")),
        _normalize(str(_first(row, "keyword", "키워드", "검색어") or "")),
    )


def _add_metrics(base: dict[str, Any], extra: dict[str, Any]) -> None:
    base["rank"] = _weighted_rank(base, extra)
    for field in ("impressions", "clicks", "cost", "conversions"):
        base[field] = _to_number(_first(base, field, *_aliases_for_field(field))) + _to_number(_first(extra, field, *_aliases_for_field(field)))


def _weighted_rank(base: dict[str, Any], extra: dict[str, Any]) -> float:
    base_rank = _to_number(_first(base, "rank", *_aliases_for_field("rank")))
    extra_rank = _to_number(_first(extra, "rank", *_aliases_for_field("rank")))
    base_impressions = _to_number(_first(base, "impressions", *_aliases_for_field("impressions")))
    extra_impressions = _to_number(_first(extra, "impressions", *_aliases_for_field("impressions")))
    total = base_impressions + extra_impressions
    if total:
        return ((base_rank * base_impressions) + (extra_rank * extra_impressions)) / total
    return base_rank or extra_rank


def _recalculate_metrics(row: dict[str, Any]) -> dict[str, Any]:
    clicks = _to_number(_first(row, "clicks", *_aliases_for_field("clicks")))
    impressions = _to_number(_first(row, "impressions", *_aliases_for_field("impressions")))
    cost = _to_number(_first(row, "cost", *_aliases_for_field("cost")))
    conversions = _to_number(_first(row, "conversions", *_aliases_for_field("conversions")))
    row["ctr"] = clicks / impressions if impressions else 0
    row["cpc"] = cost / clicks if clicks else 0
    row["conversion_rate"] = conversions / clicks if clicks else 0
    row["cost_per_conversion"] = cost / conversions if conversions else 0
    return row


def _sheet_dedupe_key(row: dict[str, Any]) -> tuple[str, ...]:
    prepared = _prepare_write_row(row)
    return tuple(_dedupe_value(prepared.get(field)) for field in HEADER_ALIASES)


def _dedupe_value(value: Any) -> str:
    if isinstance(value, float):
        return f"{value:.12g}"
    return str(value or "").strip()


def _build_sheet_targets_for_rows(rule: BrandUploadRule, current: date, rows: list[dict[str, Any]]) -> list[SheetTarget]:
    targets = build_sheet_targets(rule, current)
    if rule.mode != "fixed_today_offset" or any(_rows_for_target(rows, target) for target in targets):
        return targets

    row_dates = {_parse_row_date(_first(row, "date", "날짜", "일자")) for row in rows}
    if current in row_dates:
        direct_date_rule = replace(rule, today_offset=0, use_today_formula=False)
        return build_sheet_targets(direct_date_rule, current)
    return targets


def _row_matches_target(row: dict[str, Any], target: SheetTarget) -> bool:
    sheet_name = str(_first(row, "sheet_name", "시트명") or "").strip()
    if sheet_name:
        return _normalize(sheet_name) == _normalize(target.sheet_name)

    row_date = _parse_row_date(_first(row, "date", "날짜", "일자"))
    if not row_date:
        return False
    if row_date != target.report_date:
        return False

    media = _normalize(str(_first(row, "media", "매체") or ""))
    if target.media == "google" and "google" not in media and "구글" not in media:
        return False
    if target.media != "google" and ("google" in media or "구글" in media):
        return False

    if not _matches_campaign_kind(row, target):
        return False

    category = _normalize(str(_first(row, "category", "카테고리") or ""))
    target_category = _normalize(_base_target_category(target.category))
    if not category:
        return False
    if target.brand in {"법무법인 태하", "법무법인 오현"}:
        return category == target_category
    return category in target_category or target_category in category


def _has_positive_cost(row: dict[str, Any]) -> bool:
    return _to_number(_first(row, "cost", "총비용")) > 0


def _base_target_category(category: str) -> str:
    text = str(category)
    if text.startswith("파컨(") and text.endswith(")"):
        return text[len("파컨("):-1]
    for prefix in ("ohcrime(", "ohdcrime(", "ohehon(", "ohscrime("):
        if text.startswith(prefix) and text.endswith(")"):
            return text[len(prefix):-1]
    return text


def _matches_campaign_kind(row: dict[str, Any], target: SheetTarget) -> bool:
    if not target.campaign_kind:
        return True
    campaign_type = _normalize(str(_first(row, "campaign_type", "캠페인유형") or ""))
    is_power_contents = any(token in campaign_type for token in ("파워컨텐츠", "파워콘텐츠", "powercontents", "powercontent"))
    is_power_link = "파워링크" in campaign_type or "powerlink" in campaign_type
    if target.campaign_kind == "power_contents":
        return is_power_contents
    if target.campaign_kind == "power_link":
        return is_power_link and not is_power_contents
    return True


def _find_header_row_openpyxl(worksheet) -> int | None:
    for row_idx in range(1, min(worksheet.max_row, 12) + 1):
        values = {str(cell.value or "").strip() for cell in worksheet[row_idx]}
        normalized_values = {_normalize_header_label(value) for value in values}
        if (
            _has_any_header(normalized_values, "keyword")
            and (_has_any_header(normalized_values, "cost") or _has_any_header(normalized_values, "rank"))
        ):
            return row_idx
    return None


def _header_map_openpyxl(worksheet, header_row: int) -> dict[str, int]:
    headers = {str(cell.value or "").strip(): cell.column for cell in worksheet[header_row]}
    normalized_headers = {_normalize_header_label(name): column for name, column in headers.items()}
    result: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in headers:
                result[field] = headers[alias]
                break
            normalized = _normalize_header_label(alias)
            if normalized in normalized_headers:
                result[field] = normalized_headers[normalized]
                break
    return result


def _first_data_sheet_name(workbook) -> str | None:
    for worksheet in workbook.worksheets:
        if _find_header_row_openpyxl(worksheet) is not None:
            return worksheet.title
    return workbook.sheetnames[0] if workbook.sheetnames else None


def _has_any_header(normalized_values: set[str], field: str) -> bool:
    return any(_normalize_header_label(alias) in normalized_values for alias in HEADER_ALIASES.get(field, ()))


def _normalize_header_label(value: Any) -> str:
    return re.sub(r"[\s_()/%\[\]\-]+", "", str(value or "").casefold())


def _clear_openpyxl_data(worksheet, start_row: int, columns: Iterable[int]) -> None:
    for row_idx in range(start_row, worksheet.max_row + 1):
        for col_idx in columns:
            worksheet.cell(row_idx, col_idx).value = None


def _append_start_row_openpyxl(worksheet, start_row: int, header_map: dict[str, int]) -> int:
    key_columns = [header_map[field] for field in ("keyword", "cost") if field in header_map]
    if not key_columns:
        key_columns = list(header_map.values())
    last_data_row = start_row - 1
    for row_idx in range(start_row, worksheet.max_row + 1):
        if any(worksheet.cell(row_idx, col_idx).value not in (None, "") for col_idx in key_columns):
            last_data_row = row_idx
    return last_data_row + 1


def _value_for_field(row: dict[str, Any], field: str) -> Any:
    if field == "date":
        return _date_display_value(_first(row, field, *_aliases_for_field(field)))
    if field in {"impressions", "clicks", "cost", "conversions", "rank"}:
        return _to_number(_first(row, field, *_aliases_for_field(field)))
    if field == "ctr":
        return _calc_or_existing(row, "ctr", "클릭률", numerator="clicks", denominator="impressions")
    if field == "cpc":
        return _calc_or_existing(row, "cpc", "클릭비용", numerator="cost", denominator="clicks")
    if field == "conversion_rate":
        return _rate_or_calc(row, "conversion_rate", "전환율", numerator="conversions", denominator="clicks")
    if field == "cost_per_conversion":
        return _rate_or_calc(row, "cost_per_conversion", "전환당비용", numerator="cost", denominator="conversions")
    return _first(row, field, *_aliases_for_field(field))


def _date_display_value(value: Any) -> str:
    parsed = _parse_row_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed else str(value or "")


def _prepare_write_row(row: dict[str, Any]) -> dict[str, Any]:
    prepared = dict(row)
    for field in HEADER_ALIASES:
        prepared[field] = _value_for_field(row, field)
    _force_zero_fields(prepared)
    return prepared


def _force_zero_fields(row: dict[str, Any]) -> None:
    for field in ("conversions", "conversion_rate", "cost_per_conversion"):
        if row.get(field) in (None, ""):
            row[field] = 0
        else:
            row[field] = _to_number(row[field])


def _rate_or_calc(row: dict[str, Any], *keys: str, numerator: str, denominator: str) -> Any:
    existing = _first(row, *keys)
    if existing not in (None, ""):
        return _to_number(existing)
    n = _to_number(_first(row, numerator, *_aliases_for_field(numerator)))
    d = _to_number(_first(row, denominator, *_aliases_for_field(denominator)))
    return n / d if d else 0


def _calc_or_existing(row: dict[str, Any], *keys: str, numerator: str, denominator: str) -> Any:
    n = _to_number(_first(row, numerator, *_aliases_for_field(numerator)))
    d = _to_number(_first(row, denominator, *_aliases_for_field(denominator)))
    if d:
        return n / d
    existing = _first(row, *keys)
    return _to_number(existing) if existing not in (None, "") else 0


def _aliases_for_field(field: str) -> tuple[str, ...]:
    return HEADER_ALIASES.get(field, ())


def _first(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return ""


def _parse_row_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    match = re.search(r"(\d{4})[.\-/년 ]+(\d{1,2})[.\-/월 ]+(\d{1,2})", text)
    if match:
        y, m, d = match.groups()
        return date(int(y), int(m), int(d))
    return None


def _normalize(value: str) -> str:
    return "".join(str(value).lower().replace("_", "").replace(" ", "").split())


def _resolve_workbook_sheet_name(sheet_names: Iterable[str], target_name: str) -> str | None:
    exact = str(target_name)
    for sheet_name in sheet_names:
        if sheet_name == exact:
            return sheet_name
    normalized_target = _normalize(exact)
    for sheet_name in sheet_names:
        if _normalize(sheet_name) == normalized_target:
            return sheet_name
    return None


def _to_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").replace(",", "").replace("%", "").strip()
    if text in ("", "-"):
        return 0.0
    try:
        return float(text) / 100 if "%" in str(value) else float(text)
    except ValueError:
        return 0.0


def _save_openpyxl_workbook(workbook: Any, output: Path) -> None:
    _request_formula_recalculation(workbook)
    try:
        workbook.save(output)
    except OSError:
        with tempfile.NamedTemporaryFile(suffix=output.suffix, delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            workbook.save(tmp_path)
            _move_with_powershell(tmp_path, output)
        finally:
            try:
                tmp_path.unlink()
            except OSError:
                pass


def _ensure_directory(path: Path) -> None:
    try:
        path.mkdir(parents=True, exist_ok=True)
        return
    except OSError:
        pass
    env = os.environ.copy()
    env["BRAND_UPLOAD_DIR"] = os.fspath(path)
    result = subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", "[System.IO.Directory]::CreateDirectory($env:BRAND_UPLOAD_DIR) | Out-Null"],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
        env=env,
    )
    if result.returncode != 0:
        raise OSError((result.stderr or result.stdout or f"directory create failed: {path}").strip())


def _request_formula_recalculation(workbook: Any) -> None:
    calculation = getattr(workbook, "calculation", None) or getattr(workbook, "calculation_properties", None)
    if calculation is None:
        return
    for attr, value in (
        ("calcMode", "auto"),
        ("fullCalcOnLoad", True),
        ("forceFullCalc", True),
    ):
        try:
            setattr(calculation, attr, value)
        except Exception:
            pass


def _move_with_powershell(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    command = (
        "Move-Item -LiteralPath $env:BRAND_UPLOAD_SRC "
        "-Destination $env:BRAND_UPLOAD_DST -Force"
    )
    env = os.environ.copy()
    env["BRAND_UPLOAD_SRC"] = os.fspath(src)
    env["BRAND_UPLOAD_DST"] = os.fspath(dst)
    result = subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", command],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
        env=env,
    )
    if result.returncode != 0:
        raise OSError((result.stderr or result.stdout or "PowerShell Move-Item failed").strip())


_POWERSHELL_XLSB_WRITER = r'''
$payloadPath = [Console]::In.ReadToEnd().Trim()
$payload = Get-Content -LiteralPath $payloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$excel.ScreenUpdating = $false
$excel.EnableEvents = $false
$previousCalculation = $null
try { $previousCalculation = $excel.Calculation; $excel.Calculation = -4135 } catch {}
function Get-RowValue($row, $field, $aliases) {
  $prop = $row.PSObject.Properties[$field]
  if ($null -ne $prop -and (Test-HasExcelValue $prop.Value)) { return Convert-ExcelValue $prop.Value }
  foreach ($alias in $aliases) {
    $aliasProp = $row.PSObject.Properties[[string]$alias]
    if ($null -ne $aliasProp -and (Test-HasExcelValue $aliasProp.Value)) { return Convert-ExcelValue $aliasProp.Value }
  }
  return $null
}
function Test-HasExcelValue($value) {
  if ($null -eq $value) { return $false }
  if ($value -is [string]) { return $value -ne '' }
  return $true
}
function Convert-ExcelValue($value) {
  if ($null -eq $value) { return $null }
  if ($value -is [decimal]) { return [double]$value }
  if ($value -is [int] -or $value -is [long] -or $value -is [single] -or $value -is [double]) { return [double]$value }
  return [string]$value
}
function Normalize-SheetName($value) {
  if ($null -eq $value) { return '' }
  return ([string]$value).ToLowerInvariant().Replace('_', '').Replace(' ', '').Trim()
}
function Normalize-HeaderName($value) {
  if ($null -eq $value) { return '' }
  return ([regex]::Replace(([string]$value).ToLowerInvariant(), '[\s_()/%\[\]\-]+', '')).Trim()
}
function Resolve-Worksheet($workbook, $targetName) {
  if ($targetName -eq '__SINGLE_SHEET__') {
    foreach ($sheet in $workbook.Worksheets) { return $sheet }
  }
  foreach ($sheet in $workbook.Worksheets) {
    if ($sheet.Name -eq $targetName) { return $sheet }
  }
  $normalizedTarget = Normalize-SheetName $targetName
  foreach ($sheet in $workbook.Worksheets) {
    if ((Normalize-SheetName $sheet.Name) -eq $normalizedTarget) { return $sheet }
  }
  return $null
}
try {
  $wb = $excel.Workbooks.Open($payload.template_path, $null, $false)
  foreach ($target in $payload.targets) {
    $ws = Resolve-Worksheet $wb $target.sheet_name
    if ($null -eq $ws) { continue }
    $used = $ws.UsedRange
    $lastUsedRow = $used.Row + $used.Rows.Count - 1
    $lastUsedCol = $used.Column + $used.Columns.Count - 1
    $headerRow = 0
    for ($r = $used.Row; $r -le [Math]::Min($lastUsedRow, $used.Row + 14); $r++) {
      $vals = @()
      for ($c = $used.Column; $c -le $lastUsedCol; $c++) { $vals += (Normalize-HeaderName ([string]$ws.Cells.Item($r,$c).Text)) }
      if (($vals -contains (Normalize-HeaderName '키워드')) -and (($vals -contains (Normalize-HeaderName '총비용')) -or ($vals -contains (Normalize-HeaderName '노출순위')) -or ($vals -contains (Normalize-HeaderName 'cost')))) {
        $headerRow = $r
        break
      }
    }
    if ($headerRow -eq 0) { continue }
    $headerMap = @{}
    $normalizedHeaderMap = @{}
    for ($c = $used.Column; $c -le $lastUsedCol; $c++) {
      $name = [string]$ws.Cells.Item($headerRow,$c).Text
      if ($name) {
        $headerMap[$name] = $c
        $normalizedHeaderMap[(Normalize-HeaderName $name)] = $c
      }
    }
    $fieldCols = @{}
    foreach ($field in $payload.headers.PSObject.Properties.Name) {
      foreach ($alias in $payload.headers.$field) {
        if ($headerMap.ContainsKey($alias)) { $fieldCols[$field] = $headerMap[$alias]; break }
        $normalizedAlias = Normalize-HeaderName $alias
        if ($normalizedHeaderMap.ContainsKey($normalizedAlias)) { $fieldCols[$field] = $normalizedHeaderMap[$normalizedAlias]; break }
      }
    }
    $dataStart = $headerRow + 2
    $writeStart = $dataStart
    if ($payload.update_mode -eq 'append') {
      $keyCols = @()
      foreach ($keyField in @('keyword', 'cost')) {
        if ($fieldCols.ContainsKey($keyField)) { $keyCols += [int]$fieldCols[$keyField] }
      }
      if ($keyCols.Count -eq 0) { $keyCols = @($fieldCols.Values | ForEach-Object { [int]$_ }) }
      $lastDataRow = $dataStart - 1
      for ($r = $dataStart; $r -le $lastUsedRow; $r++) {
        foreach ($col in $keyCols) {
          if ([string]$ws.Cells.Item($r, [int]$col).Text -ne '') {
            $lastDataRow = $r
            break
          }
        }
      }
      $writeStart = $lastDataRow + 1
    }
    $rowCount = @($target.rows).Count
    $lastClearRow = [Math]::Max($lastUsedRow, $dataStart + $rowCount + 50)
    if ($rowCount -gt 0) {
      $formatCols = @($fieldCols.Values | ForEach-Object { [int]$_ } | Sort-Object)
      if ($formatCols.Count -gt 0) {
        $firstFormatCol = $formatCols[0]
        $lastFormatCol = $formatCols[$formatCols.Count - 1]
        $sourceFormat = $ws.Range($ws.Cells.Item($dataStart, $firstFormatCol), $ws.Cells.Item($dataStart, $lastFormatCol))
        $targetFormat = $ws.Range($ws.Cells.Item($writeStart, $firstFormatCol), $ws.Cells.Item([Math]::Max($lastClearRow, $writeStart + $rowCount - 1), $lastFormatCol))
        $sourceFormat.Copy() | Out-Null
        $targetFormat.PasteSpecial(-4122) | Out-Null
        try { $excel.CutCopyMode = 0 } catch {}
      }
    }
    if ($payload.update_mode -ne 'append') {
      foreach ($col in $fieldCols.Values) {
        $ws.Range($ws.Cells.Item($dataStart, [int]$col), $ws.Cells.Item($lastClearRow, [int]$col)).ClearContents() | Out-Null
      }
    }
    if ($rowCount -eq 0) { continue }
    foreach ($field in $fieldCols.Keys) {
      $col = [int]$fieldCols[$field]
      $values = New-Object 'object[,]' $rowCount, 1
      for ($i = 0; $i -lt $rowCount; $i++) {
        $value = Get-RowValue $target.rows[$i] $field $payload.headers.$field
        if (($field -eq 'conversions' -or $field -eq 'conversion_rate' -or $field -eq 'cost_per_conversion') -and $null -eq $value) { $value = 0 }
        $values[$i, 0] = $value
      }
      $writeRange = $ws.Range($ws.Cells.Item($writeStart, $col), $ws.Cells.Item($writeStart + $rowCount - 1, $col))
      $writeRange.Value2 = $values
      if ($field -eq 'conversions') { $writeRange.NumberFormat = '0;-0;-' }
      if ($field -eq 'conversion_rate') { $writeRange.NumberFormat = '0.00%;-0.00%;-' }
      if ($field -eq 'cost_per_conversion') { $writeRange.NumberFormat = '#,##0;-#,##0;-' }
    }
  }
  try { $wb.ForceFullCalculation = $true } catch {}
  try { $excel.CalculateFullRebuild() } catch { try { $wb.Calculate() } catch {} }
  $wb.SaveAs($payload.output_path, 50)
  $wb.Close($false)
}
finally {
  try { if ($null -ne $previousCalculation) { $excel.Calculation = $previousCalculation } } catch {}
  try { $excel.EnableEvents = $true } catch {}
  try { $excel.ScreenUpdating = $true } catch {}
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}
'''
