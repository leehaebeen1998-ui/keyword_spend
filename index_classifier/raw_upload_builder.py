from __future__ import annotations

import csv
import io
import json
import os
import subprocess
import tempfile
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from .brand_settings import resolve_app_path
from .exchange_rate import needs_conversion


UPLOAD_FIELDS: tuple[str, ...] = (
    "date",
    "media",
    "category",
    "campaign_type",
    "device",
    "keyword",
    "impressions",
    "clicks",
    "ctr",
    "cpc",
    "cost",
    "conversions",
    "conversion_rate",
    "cost_per_conversion",
    "rank",
)


OHYUN_URL_CATEGORY: dict[str, str] = {
    "ohcrime.com": "형사",
    "ohdcrime.com": "마약",
    "ohehon.com": "이혼",
    "ohscrime.com": "성범죄",
    "법무법인오현.com": "법무법인오현(허브)",
    "경제범죄전문변호사.com": "경제범죄",
    "사기죄전문변호사.com": "경제범죄",
    "사기죄전문변호사.kr": "경제범죄",
    "부동산전문변호사.com": "부동산전문",
    "명예훼손전문변호사.com": "명예훼손",
    "교통사고전문변호사.kr": "교통사고",
    "개인회생.org": "회생",
    "ohlabor.com": "노무노사",
    "학교폭력전문변호사.com": "학교폭력",
    "학폭변호사.kr": "학교폭력",
    "군형사전문변호사.com": "군형사",
}


TAEHA_CATEGORY_PATTERNS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("성범죄", ("성범죄", "성전문", "성폭력", "강제추행", "준강간", "카촬", "몰카", "추행", "성매매", "아청", "공연음란")),
    ("마", ("마약", "대마", "필로폰", "향정", "마약류")),
    ("이", ("이혼", "상간", "양육권", "재산분할", "위자료", "친권", "상속")),
    ("교통", ("교통", "음주", "운전", "뺑소니", "보복운전", "면허", "중과실", "사고")),
    ("군형사", ("군형사", "군사", "군대", "군인", "군변호사")),
    ("조세", ("조세", "세무", "탈세", "관세", "국세", "세금")),
    ("행정", ("행정", "영업정지", "면허취소", "소청", "징계", "학교폭력", "학폭")),
    ("개인회생", ("개인회생", "회생", "파산", "채권추심")),
    ("재산범죄", ("사기", "횡령", "배임", "보이스피싱", "전세사기", "재산범죄", "유사수신")),
    ("형", ("형사", "변호사", "법률상담", "고소", "고발", "구속", "폭행", "상해", "협박", "명예훼손", "모욕")),
)

TAEHA_NAVER_ACCOUNT_CATEGORY: dict[str, str] = {
    "01": "형",
    "02": "마",
    "03": "이",
    "05": "성",
}

TAEHA_NAVER_ACCOUNT04_PATTERNS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("조세", ("조세", "세무", "탈세", "관세", "국세", "세금", "조세범", "조세포탈")),
    ("군형사", ("군형사", "군사", "군대", "군인", "군변호사", "군전문")),
    ("행정", ("행정", "영업정지", "면허취소", "소청", "징계", "학교폭력", "학폭")),
)

UPLOAD_RULE_HEADERS: tuple[str, ...] = (
    "브랜드",
    "매체",
    "순위",
    "규칙",
    "매칭값",
    "카테고리",
    "신뢰도",
    "사용",
    "메모",
)

UPLOAD_RULE_FIELD_BY_LABEL: dict[str, str] = {
    "계정번호": "account",
    "지정 URL": "url",
    "캠페인명": "campaign",
}


@dataclass(frozen=True)
class RawBuildResult:
    output_path: Path
    total_rows: int
    written_rows: int
    skipped_rows: int
    category_counts: dict[str, int]


def build_upload_csv_from_raw(
    *,
    brand: str,
    media: str,
    input_path: str | Path,
    output_path: str | Path,
    default_category: str = "",
    rules_path: str | Path | None = None,
    exchange_rate: float | None = None,
) -> RawBuildResult:
    rows = build_upload_rows_from_raw(
        brand=brand,
        media=media,
        input_path=input_path,
        default_category=default_category,
        rules_path=rules_path,
        exchange_rate=exchange_rate,
    )
    rows = sort_upload_rows(rows)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    file = io.StringIO(newline="")
    writer = csv.DictWriter(file, fieldnames=list(UPLOAD_FIELDS))
    writer.writeheader()
    writer.writerows(rows)
    _write_text_via_temp_move(output, file.getvalue(), encoding="utf-8-sig")
    counts: dict[str, int] = {}
    for row in rows:
        category = str(row.get("category") or "")
        counts[category] = counts.get(category, 0) + 1
    return RawBuildResult(
        output_path=output,
        total_rows=len(rows),
        written_rows=len(rows),
        skipped_rows=0,
        category_counts=counts,
    )


def build_upload_rows_from_raw(
    *,
    brand: str,
    media: str,
    input_path: str | Path,
    default_category: str = "",
    rules_path: str | Path | None = None,
    exchange_rate: float | None = None,
) -> list[dict[str, Any]]:
    path = Path(input_path)
    rows, headers = _read_raw_table(path)
    return build_upload_rows_from_records(
        brand=brand,
        media=media,
        rows=rows,
        source_path=path,
        default_category=default_category,
        rules_path=rules_path,
        exchange_rate=exchange_rate,
    )


def build_upload_rows_from_records(
    *,
    brand: str,
    media: str,
    rows: list[dict[str, Any]],
    source_path: str | Path,
    default_category: str = "",
    rules_path: str | Path | None = None,
    exchange_rate: float | None = None,
) -> list[dict[str, Any]]:
    path = Path(source_path)
    external_rules = _load_brand_rules(rules_path, brand=brand) if rules_path else None
    if not external_rules:
        raise ValueError("rules_path is required. 브랜드별 업로드 규칙 CSV/JSON을 선택해 주세요.")
    return [
        _build_generic_row(row, media=media, path=path, rules=external_rules, exchange_rate=exchange_rate)
        for row in rows
        if not _skip_raw_row_for_file(row, path=path)
    ]


def build_ilo_merged_upload_rows(
    *,
    brand: str,
    media: str,
    keyword_path: str | Path,
    search_path: str | Path,
    rules_path: str | Path | None = None,
    exchange_rate: float | None = None,
) -> list[dict[str, Any]]:
    keyword_rows, _ = _read_raw_table(Path(keyword_path))
    search_rows, _ = _read_raw_table(Path(search_path))
    merged: list[dict[str, Any]] = []

    for row in search_rows:
        if str(row.get("검색 유형") or "").strip() == "일치":
            continue
        merged.append(dict(row))

    for row in keyword_rows:
        if str(row.get("검색 유형") or "").strip() != "일치":
            continue
        converted = dict(row)
        converted["검색어"] = converted.get("키워드", "")
        converted.pop("키워드", None)
        converted["검색 유형"] = "확장"
        merged.append(converted)

    return build_upload_rows_from_records(
        brand=brand,
        media=media,
        rows=merged,
        source_path=Path(search_path),
        rules_path=rules_path,
        exchange_rate=exchange_rate,
    )


def sort_upload_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: _clean_number(row.get("cost")), reverse=True)


def _skip_raw_row_for_file(row: dict[str, Any], *, path: Path) -> bool:
    campaign_type = str(row.get("캠페인유형") or row.get("캠페인 유형") or row.get("campaign_type") or "")
    campaign = str(row.get("캠페인") or row.get("캠페인명") or row.get("campaign") or "")
    if "스프레드" in path.name:
        return not _is_powercontents_text(campaign_type)
    if _is_taeha_naver_raw(path):
        return False
    return _is_powercontents_text(campaign_type) and not campaign.strip()


def _is_taeha_naver_raw(path: Path) -> bool:
    text = path.as_posix().casefold()
    name = path.name.casefold()
    return "naver" in text and ("thlaw" in name or "태하" in text)


def _load_brand_rules(rules_path: str | Path, *, brand: str) -> dict[str, Any] | None:
    path = resolve_app_path(rules_path)
    if path.suffix.lower() == ".csv":
        return _load_brand_rules_from_csv(path, brand=brand)
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    brands = data.get("brands", data)
    return brands.get(brand)


def _load_brand_rules_from_csv(path: Path, *, brand: str) -> dict[str, Any] | None:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        rows = [dict(row) for row in csv.DictReader(file)]

    rules: list[dict[str, Any]] = []
    for row in rows:
        if str(row.get("사용") or "O").strip().upper() not in ("O", "Y", "TRUE", "1"):
            continue
        row_brand = str(row.get("브랜드") or "").strip()
        if row_brand and row_brand != brand:
            continue
        label = str(row.get("규칙") or "").strip()
        field = UPLOAD_RULE_FIELD_BY_LABEL.get(label)
        match_value = str(row.get("매칭값") or "").strip()
        category = str(row.get("카테고리") or "").strip()
        if not field or not match_value or not category:
            continue
        rules.append(
            {
                "field": field,
                "match": "equals" if field == "account" else "contains",
                "patterns": [match_value],
                "category": category,
                "media": str(row.get("매체") or "").strip(),
                "priority": _safe_int(row.get("순위"), default=99),
            }
        )

    rules.sort(key=lambda item: int(item.get("priority", 99)))
    return {
        "account_regex": r"(?:^|_)(?:thlaw_|법무법인태하_)?(\d{2})(?:_|$)",
        "default_category": "미분류",
        "rules": rules,
    }


def _build_generic_row(
    row: dict[str, Any],
    *,
    media: str,
    path: Path,
    rules: dict[str, Any],
    exchange_rate: float | None = None,
) -> dict[str, Any]:
    fields = rules.get("fields", {})
    date_value = _field(row, fields, "date", "일별", "일", "date")
    campaign_type = _field(row, fields, "campaign_type", "캠페인유형", "캠페인 유형", "campaign_type")
    campaign = _field(row, fields, "campaign", "캠페인", "캠페인명", "campaign")
    device = _field(row, fields, "device", "PC/모바일 매체", "기기", "device")
    keyword = _field(row, fields, "keyword", "키워드", "검색어", "검색 키워드", "keyword")
    keyword = _keyword_for_upload(keyword=keyword, campaign=campaign, campaign_type=campaign_type)
    url = _field(row, fields, "url", "URL", "url")

    context = {
        "file_name": path.name,
        "file_stem": path.stem,
        "account": _infer_account_from_rules(path, rules, media=media),
        "url": str(url),
        "campaign": str(campaign),
        "campaign_type": str(campaign_type),
        "keyword": str(keyword),
        "media": media,
    }
    category = _category_from_rules(context, rules)

    impressions = _clean_number(_field(row, fields, "impressions", "노출수", "impressions"))
    clicks = _clean_number(_field(row, fields, "clicks", "클릭수", "clicks"))

    currency_code = _field(row, fields, "currency", "통화 코드", "통화", "currency_code")
    cost = _clean_number(_field(row, fields, "cost", "총비용", "비용", "cost"))
    cost_per_conversion = _clean_number(
        _field(row, fields, "cost_per_conversion", "총 전환당비용(원)", "모든 전환당 비용", "cost_per_conversion")
    )
    if exchange_rate and needs_conversion(currency_code):
        cost = _round2(cost * exchange_rate)
        cost_per_conversion = _round2(cost_per_conversion * exchange_rate)

    return {
        "date": _date_to_yyyymmdd(date_value) or _infer_report_date_from_path(path),
        "media": media,
        "category": category,
        "campaign_type": campaign_type,
        "device": _device(device),
        "keyword": keyword,
        "impressions": impressions,
        "clicks": clicks,
        "ctr": _calc_rate_or_percent(clicks, impressions, _field(row, fields, "ctr", "클릭률(%)", "클릭률(CTR)", "ctr")),
        "cpc": _safe_div(cost, clicks),
        "cost": cost,
        "conversions": _clean_number(_field(row, fields, "conversions", "총 전환수", "전환", "conversions")),
        "conversion_rate": _percent(_field(row, fields, "conversion_rate", "총 전환율(%)", "전환율", "conversion_rate")),
        "cost_per_conversion": cost_per_conversion,
        "rank": _clean_number(_field(row, fields, "rank", "평균노출순위", "rank")),
    }


def _field(row: dict[str, Any], fields: dict[str, str], name: str, *fallbacks: str) -> Any:
    normalized = {_normalize_header(key): key for key in row}
    configured = fields.get(name)
    if configured:
        actual = normalized.get(_normalize_header(configured))
        if actual is not None:
            return row.get(actual, "")
    for fallback in fallbacks:
        actual = normalized.get(_normalize_header(fallback))
        if actual is not None:
            return row.get(actual, "")
    return ""


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s_()/%\[\]\-]+", "", str(value or "").casefold())


def _keyword_for_upload(*, keyword: Any, campaign: Any, campaign_type: Any) -> Any:
    keyword_text = str(keyword or "").strip()
    campaign_text = str(campaign or "").strip()
    campaign_type_text = str(campaign_type or "").strip()
    is_power_contents = "파워컨텐츠" in campaign_text or "파워콘텐츠" in campaign_text or "파워컨텐츠" in campaign_type_text or "파워콘텐츠" in campaign_type_text
    if is_power_contents and campaign_text and keyword_text in ("", "-"):
        return campaign_text
    return keyword


def _safe_int(value: Any, *, default: int) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _safe_div(numerator: Any, denominator: Any) -> float:
    n = _clean_number(numerator)
    d = _clean_number(denominator)
    return n / d if d else 0.0


def _calc_rate_or_percent(numerator: Any, denominator: Any, fallback: Any = "") -> float:
    result = _safe_div(numerator, denominator)
    return result if result else _percent(fallback)


def _infer_report_date_from_path(path: Path) -> str:
    text = path.name
    match = re.search(r"_raw_(20\d{6})(?:_|$)", text)
    if match:
        return match.group(1)
    match = re.search(r"(20\d{6})", text)
    return match.group(1) if match else ""


def _infer_account_from_rules(path: Path, rules: dict[str, Any], *, media: str) -> str:
    if "google" in media.lower() or "구글" in media:
        return ""
    account_regex = rules.get("account_regex")
    if account_regex:
        match = re.search(str(account_regex), path.stem)
        if match:
            for group in match.groups():
                if group:
                    return group
    for pattern in (r",(\d{4,})(?:$|_)", r"_(\d{4,})(?:_|$)", r"(\d{4,})"):
        match = re.search(pattern, path.stem)
        if match:
            return match.group(1)
    return ""


def _category_from_rules(context: dict[str, str], rules: dict[str, Any]) -> str:
    for rule in rules.get("rules", []):
        field = str(rule.get("field") or "")
        category = str(rule.get("category") or "")
        if not field or not category:
            continue
        if not _media_matches(context.get("media", ""), str(rule.get("media") or "")):
            continue
        value = context.get(field, "")
        if _rule_matches(str(value), rule):
            return category
        if field == "campaign" and _is_powercontents_text(context.get("keyword", "")) and _rule_matches(context.get("keyword", ""), rule):
            return category
    return str(rules.get("default_category") or "미분류")


def _media_matches(row_media: str, rule_media: str) -> bool:
    if not rule_media or rule_media in ("전체", "ALL", "all"):
        return True
    return _normalize_media(row_media) == _normalize_media(rule_media)


def _normalize_media(value: str) -> str:
    normalized = str(value).casefold().replace(" ", "").replace("_", "")
    if "google" in normalized or "구글" in normalized:
        return "google"
    if "naver" in normalized or "네이버" in normalized:
        return "naver"
    return normalized


def _rule_matches(value: str, rule: dict[str, Any]) -> bool:
    match_type = str(rule.get("match", "contains"))
    patterns = rule.get("patterns", [])
    if isinstance(patterns, str):
        patterns = [patterns]
    normalized_value = value.lower()
    for pattern in patterns:
        text = str(pattern)
        if match_type == "equals" and normalized_value == text.lower():
            return True
        if match_type == "contains" and text.lower() in normalized_value:
            return True
        if match_type == "regex" and re.search(text, value):
            return True
    return False


def _is_powercontents_text(value: Any) -> bool:
    text = str(value or "")
    return "파워컨텐츠" in text or "파워콘텐츠" in text


def _read_raw_table(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx_table(path)
    encoding = _detect_encoding(path)
    delimiter = "\t" if encoding == "utf-16" else ","
    with path.open("r", encoding=encoding, newline="") as file:
        raw_rows = list(csv.reader(file, delimiter=delimiter))

    header_index = _detect_header_index(raw_rows)
    headers = [str(value).strip() for value in raw_rows[header_index]]
    rows: list[dict[str, Any]] = []
    for raw in raw_rows[header_index + 1:]:
        if not any(str(value).strip() for value in raw):
            continue
        rows.append({headers[idx]: raw[idx] if idx < len(raw) else "" for idx in range(len(headers))})
    return rows, headers


def _read_xlsx_table(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for spreadsheet raw upload.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    raw_rows: list[list[Any]] = []
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else value for value in row]
                if any(str(value).strip() for value in values):
                    raw_rows.append(values)
            if raw_rows:
                break
    finally:
        workbook.close()
    if not raw_rows:
        return [], []
    header_index = _detect_header_index([[str(value) for value in row] for row in raw_rows])
    headers = [str(value).strip() for value in raw_rows[header_index]]
    rows: list[dict[str, Any]] = []
    for raw in raw_rows[header_index + 1:]:
        if not any(str(value).strip() for value in raw):
            continue
        rows.append({headers[idx]: raw[idx] if idx < len(raw) else "" for idx in range(len(headers))})
    return rows, headers


def _detect_encoding(path: Path) -> str:
    raw = path.read_bytes()[:4]
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16"
    return "utf-8-sig"


def _detect_header_index(rows: list[list[str]]) -> int:
    required = {"키워드", "노출수", "클릭수"}
    for idx, row in enumerate(rows[:10]):
        values = {str(value).strip() for value in row}
        if len(values & required) >= 2:
            return idx
        if {"검색 키워드", "노출수", "클릭수"} <= values:
            return idx
    return 0


def _build_ohyun_row(row: dict[str, Any], *, media: str) -> dict[str, Any]:
    category = _ohyun_category(str(row.get("URL") or ""))
    return {
        "date": _date_to_yyyymmdd(row.get("일별")),
        "media": media,
        "category": category,
        "campaign_type": row.get("캠페인유형", ""),
        "device": row.get("PC/모바일 매체", ""),
        "keyword": row.get("키워드", ""),
        "impressions": _clean_number(row.get("노출수")),
        "clicks": _clean_number(row.get("클릭수")),
        "ctr": _calc_rate_or_percent(row.get("클릭수"), row.get("노출수"), row.get("클릭률(%)")),
        "cpc": _clean_number(row.get("평균 CPC")),
        "cost": _clean_number(row.get("총비용")),
        "conversions": _clean_number(row.get("총 전환수")),
        "conversion_rate": _percent(row.get("총 전환율(%)")),
        "cost_per_conversion": _clean_number(row.get("총 전환당비용(원)")),
        "rank": _clean_number(row.get("평균노출순위")),
    }


def _build_taeha_row(row: dict[str, Any], *, media: str, default_category: str, account_no: str) -> dict[str, Any]:
    if "검색 키워드" in row:
        campaign = str(row.get("캠페인") or "")
        keyword = str(row.get("검색 키워드") or "")
        category = _taeha_google_category(f"{campaign} {keyword}", default_category=default_category)
        return {
            "date": _date_to_yyyymmdd(row.get("일")),
            "media": media,
            "category": category,
            "campaign_type": row.get("캠페인 유형", ""),
            "device": _device(row.get("기기")),
            "keyword": keyword,
            "impressions": _clean_number(row.get("노출수")),
            "clicks": _clean_number(row.get("클릭수")),
            "ctr": _calc_rate_or_percent(row.get("클릭수"), row.get("노출수"), row.get("클릭률(CTR)")),
            "cpc": _clean_number(row.get("평균 비용")),
            "cost": _clean_number(row.get("비용")),
            "conversions": _clean_number(row.get("전환")),
            "conversion_rate": _percent(row.get("전환율")),
            "cost_per_conversion": _clean_number(row.get("모든 전환당 비용")),
            "rank": "",
        }

    keyword = str(row.get("키워드") or "")
    category = _taeha_naver_category(row, keyword=keyword, default_category=default_category, account_no=account_no)
    return {
        "date": _date_to_yyyymmdd(row.get("일별")),
        "media": media,
        "category": category,
        "campaign_type": row.get("캠페인유형", ""),
        "device": row.get("PC/모바일 매체", ""),
        "keyword": keyword,
        "impressions": _clean_number(row.get("노출수")),
        "clicks": _clean_number(row.get("클릭수")),
        "ctr": _calc_rate_or_percent(row.get("클릭수"), row.get("노출수"), row.get("클릭률(%)")),
        "cpc": _clean_number(row.get("평균 CPC")),
        "cost": _clean_number(row.get("총비용")),
        "conversions": _clean_number(row.get("총 전환수")),
        "conversion_rate": _percent(row.get("총 전환율(%)")),
        "cost_per_conversion": _clean_number(row.get("총 전환당비용(원)")),
        "rank": _clean_number(row.get("평균노출순위")),
    }


def _ohyun_category(url: str) -> str:
    normalized = url.lower().replace("https://", "").replace("http://", "").replace("www.", "")
    normalized = normalized.strip().rstrip("/")
    for key, category in OHYUN_URL_CATEGORY.items():
        if key.lower() in normalized:
            return category
    return "법무법인오현(허브)"


def _taeha_category(text: str, *, default_category: str) -> str:
    if default_category:
        return default_category
    compact = str(text).replace(" ", "")
    for category, patterns in TAEHA_CATEGORY_PATTERNS:
        if any(pattern in compact for pattern in patterns):
            return category
    return "형"


def _taeha_google_category(text: str, *, default_category: str) -> str:
    return _taeha_category(text, default_category=default_category)


def _taeha_naver_category(
    row: dict[str, Any],
    *,
    keyword: str,
    default_category: str,
    account_no: str,
) -> str:
    if default_category:
        return default_category
    if account_no in TAEHA_NAVER_ACCOUNT_CATEGORY:
        return TAEHA_NAVER_ACCOUNT_CATEGORY[account_no]
    if account_no == "04":
        text = " ".join(
            str(row.get(key) or "")
            for key in ("URL", "url", "캠페인", "캠페인명", "캠페인유형", "키워드")
        )
        compact = text.replace(" ", "")
        for category, patterns in TAEHA_NAVER_ACCOUNT04_PATTERNS:
            if any(pattern in compact for pattern in patterns):
                return category
        return "행정"
    return _taeha_category(keyword, default_category="")


def _infer_taeha_account_no(path: Path) -> str:
    stem = path.stem
    patterns = (
        r"(?:^|_)thlaw_(\d{2})(?:_|$)",
        r"(?:^|_)법무법인태하_(\d{2})(?:_|$)",
        r"(?:^|_)(\d{2})_(?:형사|마약|이혼|조세|행정|군형사|성)",
    )
    for pattern in patterns:
        match = re.search(pattern, stem)
        if match:
            return match.group(1)
    return ""


def _date_to_yyyymmdd(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.rstrip(".")
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    match = re.search(r"(\d{4})[.\-/년 ]+(\d{1,2})[.\-/월 ]+(\d{1,2})", text)
    if match:
        y, m, d = match.groups()
        return f"{int(y):04d}{int(m):02d}{int(d):02d}"
    return text


def _device(value: Any) -> str:
    text = str(value or "")
    if "휴대" in text or "모바일" in text:
        return "모바일"
    if "컴퓨터" in text or "PC" in text:
        return "PC"
    return text


def _percent(value: Any) -> float:
    text = str(value or "").replace(",", "").strip()
    if not text or text == "-":
        return 0.0
    has_percent = "%" in text
    text = text.replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return number / 100 if has_percent or number > 1 else number


def _clean_number(value: Any) -> float | int:
    text = str(value or "").replace(",", "").strip()
    if not text or text == "-":
        return 0
    try:
        number = float(text)
        return int(number) if number == int(number) else number
    except ValueError:
        return 0


def _round2(value: float | int) -> float | int:
    number = round(float(value), 2)
    return int(number) if number == int(number) else number


def _write_text_via_temp_move(path: Path, content: str, *, encoding: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        path.write_text(content, encoding=encoding, newline="")
        return
    except OSError:
        pass

    with tempfile.NamedTemporaryFile("w", encoding=encoding, newline="", delete=False, suffix=".csv") as tmp:
        tmp.write(content)
        tmp_path = Path(tmp.name)
    try:
        command = "Copy-Item -LiteralPath $env:RAW_UPLOAD_SRC -Destination $env:RAW_UPLOAD_DST -Force"
        env = os.environ.copy()
        env["RAW_UPLOAD_SRC"] = os.fspath(tmp_path)
        env["RAW_UPLOAD_DST"] = os.fspath(path)
        result = subprocess.run(
            ["powershell.exe", "-NoProfile", "-Command", command],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
            env=env,
        )
        if result.returncode != 0:
            raise OSError((result.stderr or result.stdout or "Move-Item failed").strip())
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass
